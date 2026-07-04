"""
Система учета ПУ - Backend
ЭТАП 1: Базовая структура
"""
from fastapi import FastAPI, Depends, HTTPException, UploadFile, File, Form, Query
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy import create_engine, Column, Integer, String, Boolean, DateTime, ForeignKey, Text, Enum as SQLEnum, Float, Date, or_
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import sessionmaker, Session, relationship, joinedload
from sqlalchemy.sql import func
from pydantic import BaseModel
from pydantic_settings import BaseSettings
from typing import Optional, List
from datetime import datetime, timedelta, date
from jose import jwt
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
import bcrypt as _bcrypt
import pandas as pd
import io
import json
import enum
import re
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from fastapi.responses import StreamingResponse
from urllib.parse import quote



# ==================== КОНФИГ ====================
class Settings(BaseSettings):
    DATABASE_URL: str = "postgresql://user:pass@localhost/pu_system"
    SECRET_KEY: str = "your-secret-key-change-me"
    ADMIN_CODE: str = "2233"
    class Config:
        env_file = ".env"

settings = Settings()

# ==================== БАЗА ДАННЫХ ====================
engine = create_engine(settings.DATABASE_URL)
SessionLocal = sessionmaker(bind=engine)
Base = declarative_base()

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

# ==================== ENUM'ы ====================
class UnitType(str, enum.Enum):
    SUE = "SUE"          # Служба учета электроэнергии
    LAB = "LAB"          # Лаборатория
    ESK = "ESK"          # ЭСК (центральный админ)
    RES = "RES"          # РЭС (7 штук)
    ESK_UNIT = "ESK_UNIT"  # Подразделение ЭСК (7 штук)
    OKS = "OKS"          # ОКС (центральный, видит все участки ОКС)
    OKS_UNIT = "OKS_UNIT"  # Участок ОКС (7 штук, по одному на РЭС)

class RoleCode(str, enum.Enum):
    SUE_ADMIN = "SUE_ADMIN"      # СУЭ - видит всё, перемещает РЭС, удаляет, управляет
    LAB_USER = "LAB_USER"        # Лаборатория - загружает реестры
    ESK_ADMIN = "ESK_ADMIN"      # ЭСК Админ - видит все ЭСК, перемещает между ЭСК
    RES_USER = "RES_USER"        # Пользователь РЭС - только свой РЭС
    ESK_USER = "ESK_USER"        # Пользователь ЭСК - только своё подразделение ЭСК
    OKS_ADMIN = "OKS_ADMIN"      # ОКС Админ - видит все участки ОКС, перемещает между ними
    OKS_USER = "OKS_USER"        # Пользователь ОКС - только свой участок ОКС

class PUStatus(str, enum.Enum):
    SKLAD = "SKLAD"          # На складе (по умолчанию)
    TECHPRIS = "TECHPRIS"    # Техприс
    ZAMENA = "ZAMENA"        # Замена
    IZHC = "IZHC"            # ИЖЦ
    INSTALLED = "INSTALLED"  # Установлен

class ApprovalStatus(str, enum.Enum):
    NONE = "NONE"            # Не отправлено
    PENDING = "PENDING"      # На согласовании
    APPROVED = "APPROVED"    # Согласовано
    REJECTED = "REJECTED"    # Отклонено

# ==================== МОДЕЛИ БД ====================

class Unit(Base):
    """Подразделения"""
    __tablename__ = "units"
    id = Column(Integer, primary_key=True)
    name = Column(String(200), nullable=False)
    code = Column(String(50), unique=True)
    short_code = Column(String(10))  # с, а, д, л, т, х, к для ТЗ
    unit_type = Column(SQLEnum(UnitType))
    parent_id = Column(Integer, ForeignKey("units.id"))
    is_active = Column(Boolean, default=True)

class Role(Base):
    """Роли"""
    __tablename__ = "roles"
    id = Column(Integer, primary_key=True)
    name = Column(String(100))
    code = Column(SQLEnum(RoleCode), unique=True)

class User(Base):
    """Пользователи"""
    __tablename__ = "users"
    id = Column(Integer, primary_key=True)
    username = Column(String(50), unique=True)
    password_hash = Column(String(255))
    full_name = Column(String(200))
    role_id = Column(Integer, ForeignKey("roles.id"))
    unit_id = Column(Integer, ForeignKey("units.id"))
    is_active = Column(Boolean, default=True)
    created_at = Column(DateTime, server_default=func.now())
    role = relationship("Role")
    unit = relationship("Unit")

class PURegister(Base):
    """Реестр загрузок ПУ"""
    __tablename__ = "pu_registers"
    id = Column(Integer, primary_key=True)
    filename = Column(String(255))
    uploaded_by = Column(Integer, ForeignKey("users.id"))
    uploaded_at = Column(DateTime, server_default=func.now())
    items_count = Column(Integer, default=0)
    uploader = relationship("User")

class PUItem(Base):
    """Прибор учета - расширенная карточка"""
    __tablename__ = "pu_items"
    id = Column(Integer, primary_key=True)
    register_id = Column(Integer, ForeignKey("pu_registers.id"))
    
    # Базовые поля (из импорта лаборатории)
    serial_number = Column(String(100), index=True)  # Заводской номер
    pu_type = Column(String(500))  # Тип счетчика
    
    # Местоположение
    target_unit_id = Column(Integer, ForeignKey("units.id"))  # Куда назначен
    current_unit_id = Column(Integer, ForeignKey("units.id"))  # Где сейчас
    
    # Статус и тип работы
    status = Column(SQLEnum(PUStatus), default=PUStatus.SKLAD)
    naznachenie = Column(String(20))  # Назначение из лаборатории: IZHC, TECHPRIS, ZAMENA
    
    # Поля карточки РЭС
    tz_number = Column(String(50))  # Номер ТЗ
    faza = Column(String(20))  # Фазность (из справочника)
    voltage = Column(String(20))  # Уровень напряжения 0.23, 0.4, 6, 10
    power = Column(Float)  # Мощность кВт
    
    # Для Техприс и ЭСК
    contract_number = Column(String(50))  # Договор ТП формат ххххх-хх-хххххххх-х
    contract_date = Column(Date)  # Дата заключения
    plan_date = Column(Date)  # Планируемая дата исполнения
    consumer = Column(String(500))  # Потребитель
    address = Column(Text)  # Адрес
    
    # Для Замена и ИЖЦ
    ls_number = Column(String(50))  # Лицевой счет
    
    # СМР
    smr_executor = Column(String(20))  # РСК или ЭСК
    smr_date = Column(Date)  # Дата выполнения СМР
    smr_master_id = Column(Integer, ForeignKey("esk_masters.id"))  # Мастер ЭСК
    
    # ТТР для РЭС
    ttr_ou_id = Column(Integer, ForeignKey("ttr_res.id"))  # ТТР организации учета
    ttr_ol_id = Column(Integer, ForeignKey("ttr_res.id"))  # ТТР обустройство линии
    ttr_or_id = Column(Integer, ForeignKey("ttr_res.id"))  # ТТР распред. щита
    ttr_tt_id = Column(Integer, ForeignKey("ttr_res.id"))  # ТТР для ТТ (У-27)
    
    # ТТР для ЭСК
    # Параметры СМР/ЛСР для ЭСК
    form_factor = Column(String(20))  # split, classic (автоподтяжка)
    trubostoyka = Column(Boolean, default=False)  # Да/Нет
    va_type = Column(String(20))  # opona, fasad, trubostoyka
    ttr_esk_id = Column(Integer, ForeignKey("ttr_esk.id"))
    lsr_number = Column(String(50))  # Номер ЛСР
    price_no_nds = Column(Float)  # Стоимость без НДС
    price_with_nds = Column(Float)  # Стоимость с НДС

        # ЛСР Трубостойки (отдельно)
    lsr_truba = Column(String(50))
    price_truba_no_nds = Column(Float)
    price_truba_with_nds = Column(Float)
    
    # ЛСР ВА (отдельно)
    lsr_va = Column(String(50))
    price_va_no_nds = Column(Float)
    price_va_with_nds = Column(Float)
    
    # Материалы (JSON или отдельная таблица)
    materials_used = Column(Boolean, default=False)  # Материалы использованы

    # ВА и ТТ
    has_va = Column(Boolean, default=False)
    va_nominal_id = Column(Integer, ForeignKey("va_nominals.id"))
    va_quantity = Column(Integer, default=1)
    has_tt = Column(Boolean, default=False)
    tt_nominal_id = Column(Integer, ForeignKey("tt_nominals.id"))
    
    # Согласование (для ЭСК)
    approval_status = Column(SQLEnum(ApprovalStatus), default=ApprovalStatus.NONE)
    approved_by = Column(Integer, ForeignKey("users.id"))
    approved_at = Column(DateTime)
    rejection_comment = Column(Text)
    
    # Заявка ЭСК
    request_number = Column(String(50))  # Номер заявки (например 1-26)
    request_contract = Column(String(50))  # Номер договора заявки (например 147)
    work_type_name = Column(String(200))  # Наименование вида работ (из ТТР)
    
    created_at = Column(DateTime, server_default=func.now())
    updated_at = Column(DateTime, server_default=func.now(), onupdate=func.now())
    
    # Relationships
    register = relationship("PURegister")
    target_unit = relationship("Unit", foreign_keys=[target_unit_id])
    current_unit = relationship("Unit", foreign_keys=[current_unit_id])
    ttr_ou = relationship("TTR_RES", foreign_keys=[ttr_ou_id])
    ttr_ol = relationship("TTR_RES", foreign_keys=[ttr_ol_id])
    ttr_or = relationship("TTR_RES", foreign_keys=[ttr_or_id])
    ttr_tt = relationship("TTR_RES", foreign_keys=[ttr_tt_id])
    ttr_esk = relationship("TTR_ESK", foreign_keys=[ttr_esk_id])
    va_nominal = relationship("VA_Nominal", foreign_keys=[va_nominal_id])
    tt_nominal = relationship("TT_Nominal", foreign_keys=[tt_nominal_id])

class PUMovement(Base):
    """История перемещений"""
    __tablename__ = "pu_movements"
    id = Column(Integer, primary_key=True)
    pu_item_id = Column(Integer, ForeignKey("pu_items.id"))
    from_unit_id = Column(Integer, ForeignKey("units.id"))
    to_unit_id = Column(Integer, ForeignKey("units.id"))
    moved_by = Column(Integer, ForeignKey("users.id"))
    moved_at = Column(DateTime, server_default=func.now())
    comment = Column(Text)

class TTR_RES(Base):
    """Справочник ТТР для РЭС"""
    __tablename__ = "ttr_res"
    id = Column(Integer, primary_key=True)
    code = Column(String(50))  # ТТР-1 ОУ, ТТР-2 ОЛ и т.д.
    name = Column(String(200))
    ttr_type = Column(String(20))  # OU, OL, OR (организация учета, линии, распред)
    pu_types = Column(Text)  # Для каких типов ПУ применим (JSON или через запятую)
    use_tt = Column(Boolean, default=False)
    is_active = Column(Boolean, default=True)

class TTR_ESK(Base):
    """Справочник ТТР для ЭСК (со стоимостью)"""
    __tablename__ = "ttr_esk"
    id = Column(Integer, primary_key=True)
    ttr_type = Column(String(20))  # PU, TRUBOSTOYKA, OTVETVLENIE
    work_type_name = Column(String(200))  # Наименование вида работ
    pu_pattern = Column(String(200))  # Паттерн наименования ПУ для автоопределения
    faza = Column(String(10))  # 1ф, 3ф
    form_factor = Column(String(20))  # split, classic
    va_type = Column(String(20))  # opora, fasad, trubostoyka
    lsr_number = Column(String(50))  # Номер ЛСР
    price_no_nds = Column(Float, default=0)  # Стоимость без НДС
    price_with_nds = Column(Float, default=0)  # Стоимость с НДС
    is_active = Column(Boolean, default=True)

class Material(Base):
    """Справочник материалов"""
    __tablename__ = "materials"
    id = Column(Integer, primary_key=True)
    name = Column(String(200))
    unit = Column(String(50))  # шт, м, кг
    is_active = Column(Boolean, default=True)

class TTR_Material(Base):
    """Связь ТТР и материалов (сколько чего нужно)"""
    __tablename__ = "ttr_materials"
    id = Column(Integer, primary_key=True)
    ttr_res_id = Column(Integer, ForeignKey("ttr_res.id"))
    material_id = Column(Integer, ForeignKey("materials.id"))
    quantity = Column(Float, default=0)

class TTR_PUType(Base):
    """Связь ТТР РЭС и типов ПУ"""
    __tablename__ = "ttr_pu_types"
    id = Column(Integer, primary_key=True)
    ttr_res_id = Column(Integer, ForeignKey("ttr_res.id"))
    pu_type_id = Column(Integer, ForeignKey("pu_type_reference.id"))
    ttr = relationship("TTR_RES")
    pu_type = relationship("PUTypeReference")

class PUMaterial(Base):
    """Использованные материалы в конкретном ПУ"""
    __tablename__ = "pu_materials"
    id = Column(Integer, primary_key=True)
    pu_item_id = Column(Integer, ForeignKey("pu_items.id"))
    material_id = Column(Integer, ForeignKey("materials.id"))
    quantity = Column(Float, default=0)
    used = Column(Boolean, default=True)  # Галочка использован/нет

class ESKMaster(Base):
    """Справочник мастеров ЭСК"""
    __tablename__ = "esk_masters"
    id = Column(Integer, primary_key=True)
    unit_id = Column(Integer, ForeignKey("units.id"))  # Подразделение ЭСК
    full_name = Column(String(200))
    is_active = Column(Boolean, default=True)
    unit = relationship("Unit")

class PUTypeReference(Base):
    """Справочник типов ПУ для автоопределения"""
    __tablename__ = "pu_type_reference"
    id = Column(Integer, primary_key=True)
    pattern = Column(String(200))
    faza = Column(String(20))  # 1ф, 3ф
    voltage = Column(String(20))
    form_factor = Column(String(20))  # split, classic
    is_active = Column(Boolean, default=True)


class VA_Nominal(Base):
    """Справочник номиналов ВА"""
    __tablename__ = "va_nominals"
    id = Column(Integer, primary_key=True)
    name = Column(String(100))  # Например: "16А", "25А", "32А"
    is_active = Column(Boolean, default=True)


class TT_Nominal(Base):
    """Справочник номиналов ТТ"""
    __tablename__ = "tt_nominals"
    id = Column(Integer, primary_key=True)
    name = Column(String(100))  # Например: "100/5", "200/5", "400/5"
    is_active = Column(Boolean, default=True)

# ==================== АВТОРИЗАЦИЯ ====================
security = HTTPBearer()

def hash_password(password: str) -> str:
    return _bcrypt.hashpw(password.encode(), _bcrypt.gensalt()).decode()

def verify_password(plain: str, hashed: str) -> bool:
    return _bcrypt.checkpw(plain.encode(), hashed.encode())

def create_token(user_id: int) -> str:
    expire = datetime.utcnow() + timedelta(hours=24)
    return jwt.encode({"sub": str(user_id), "exp": expire}, settings.SECRET_KEY)

def get_current_user(creds: HTTPAuthorizationCredentials = Depends(security), db: Session = Depends(get_db)) -> User:
    try:
        payload = jwt.decode(creds.credentials, settings.SECRET_KEY, algorithms=["HS256"])
        user = db.query(User).filter(User.id == int(payload["sub"])).first()
        if not user or not user.is_active:
            raise HTTPException(401, "Не авторизован")
        return user
    except:
        raise HTTPException(401, "Неверный токен")

# Проверки ролей
def is_sue_admin(user: User) -> bool:
    return user.role.code == RoleCode.SUE_ADMIN

def is_lab_user(user: User) -> bool:
    return user.role.code == RoleCode.LAB_USER

def is_esk_admin(user: User) -> bool:
    return user.role.code == RoleCode.ESK_ADMIN

def is_res_user(user: User) -> bool:
    return user.role.code == RoleCode.RES_USER

def is_esk_user(user: User) -> bool:
    return user.role.code == RoleCode.ESK_USER

def is_oks_admin(user: User) -> bool:
    return user.role.code == RoleCode.OKS_ADMIN

def is_oks_user(user: User) -> bool:
    return user.role.code == RoleCode.OKS_USER

# ==================== АВТООПРЕДЕЛЕНИЕ ТИПА ПУ ====================

def normalize_pu_string(s: str) -> str:
    """Нормализация строки ПУ: убираем невидимые символы, 
    приводим кириллицу к латинице, убираем лишние пробелы/знаки"""
    import unicodedata
    if not s:
        return ""
    
    # Убираем управляющие символы и нормализуем unicode
    s = unicodedata.normalize('NFKC', s)
    
    # Приводим к верхнему регистру
    s = s.upper()
    
    # Замена кириллических букв-двойников на латинские
    cyr_to_lat = {
        'А': 'A', 'В': 'B', 'Е': 'E', 'К': 'K', 'М': 'M',
        'Н': 'H', 'О': 'O', 'Р': 'P', 'С': 'C', 'Т': 'T',
        'У': 'Y', 'Х': 'X',
    }
    s = ''.join(cyr_to_lat.get(c, c) for c in s)
    
    # Убираем все кроме букв, цифр, точек и дефисов
    import re
    s = re.sub(r'[^\w\.\-]', ' ', s)  # заменяем спецсимволы на пробел
    s = re.sub(r'\s+', ' ', s).strip()  # схлопываем пробелы
    
    return s


def detect_pu_type_params(pu_type: str, db: Session) -> dict:
    """
    Определяет фазность, напряжение и форм-фактор по паттерну из справочника.
    Многоуровневый поиск: точное → нормализованное → токенное совпадение.
    """
    if not pu_type:
        return {}
    
    pu_norm = normalize_pu_string(pu_type)
    
    # Получаем все паттерны из справочника
    patterns = db.query(PUTypeReference).filter(PUTypeReference.is_active == True).all()
    
    # Подготавливаем паттерны с нормализацией
    prepared = []
    for p in patterns:
        if not p.pattern:
            continue
        p_norm = normalize_pu_string(p.pattern)
        if p_norm:
            prepared.append((p, p_norm))
    
    # Сортируем по длине нормализованного паттерна (длинные = точнее)
    prepared.sort(key=lambda x: len(x[1]), reverse=True)
    
    def extract_result(p):
        result = {}
        if p.faza:
            result['faza'] = p.faza
        if p.voltage:
            result['voltage'] = p.voltage
        if p.form_factor:
            result['form_factor'] = p.form_factor
        return result
    
    # 1) Точное вхождение нормализованной строки
    for p, p_norm in prepared:
        if p_norm in pu_norm or pu_norm.startswith(p_norm):
            return extract_result(p)
    
    # 2) Токенное совпадение — разбиваем на слова и ищем максимальное пересечение
    pu_tokens = set(pu_norm.split())
    
    best_match = None
    best_score = 0
    
    for p, p_norm in prepared:
        p_tokens = set(p_norm.split())
        if not p_tokens:
            continue
        
        # Сколько токенов паттерна есть в названии ПУ
        matched = len(p_tokens & pu_tokens)
        score = matched / len(p_tokens)  # доля совпавших токенов
        
        # Минимум 70% токенов должны совпасть и хотя бы 2 токена
        if score >= 0.7 and matched >= 2 and score > best_score:
            best_score = score
            best_match = p
    
    if best_match:
        return extract_result(best_match)
    
    return {}
    
def get_visible_units(user: User, db: Session) -> List[int]:
    """Какие подразделения видит пользователь"""
    if is_sue_admin(user):
        return [u.id for u in db.query(Unit).all()]
    if is_esk_admin(user):
        return [u.id for u in db.query(Unit).filter(Unit.unit_type.in_([UnitType.ESK, UnitType.ESK_UNIT])).all()]
    if is_oks_admin(user):
        return [u.id for u in db.query(Unit).filter(Unit.unit_type.in_([UnitType.OKS, UnitType.OKS_UNIT])).all()]
    if is_lab_user(user):
        return [user.unit_id] if user.unit_id else []
    # RES_USER и ESK_USER видят только своё подразделение
    return [user.unit_id] if user.unit_id else []

def can_move_pu(user: User, pu_item, target_unit, db: Session) -> tuple[bool, str]:
    """Проверка прав на перемещение"""
    if is_sue_admin(user):
        # СУЭ может перемещать только ПУ из РЭС в РЭС
        if pu_item.current_unit and pu_item.current_unit.unit_type in [UnitType.ESK, UnitType.ESK_UNIT, UnitType.OKS, UnitType.OKS_UNIT]:
            return False, "СУЭ не может перемещать ПУ из ЭСК/ОКС"
        if target_unit.unit_type in [UnitType.ESK, UnitType.ESK_UNIT, UnitType.OKS, UnitType.OKS_UNIT]:
            return False, "СУЭ может перемещать только в РЭС"
        return True, ""
    
    if is_esk_admin(user):
        # ЭСК админ может перемещать только между ЭСК
        if pu_item.current_unit and pu_item.current_unit.unit_type not in [UnitType.ESK, UnitType.ESK_UNIT]:
            return False, "ЭСК может перемещать только ПУ из ЭСК"
        if target_unit.unit_type not in [UnitType.ESK, UnitType.ESK_UNIT]:
            return False, "ЭСК может перемещать только в ЭСК"
        return True, ""
    
    if is_oks_admin(user):
        # ОКС админ может перемещать только между участками ОКС
        if pu_item.current_unit and pu_item.current_unit.unit_type not in [UnitType.OKS, UnitType.OKS_UNIT]:
            return False, "ОКС может перемещать только ПУ из ОКС"
        if target_unit.unit_type not in [UnitType.OKS, UnitType.OKS_UNIT]:
            return False, "ОКС может перемещать только в ОКС"
        return True, ""
    
    return False, "Нет прав на перемещение"

# ==================== PYDANTIC СХЕМЫ ====================
class LoginReq(BaseModel):
    username: str
    password: str

class TokenResp(BaseModel):
    access_token: str

class UserResp(BaseModel):
    id: int
    username: str
    full_name: str
    role_code: str
    role_name: str
    unit_id: Optional[int]
    unit_name: Optional[str]
    unit_type: Optional[str]
    visible_units: List[int] = []

class UnitResp(BaseModel):
    id: int
    name: str
    code: str
    unit_type: str
    short_code: Optional[str]

class MoveReq(BaseModel):
    pu_item_ids: List[int]
    to_unit_id: int
    comment: Optional[str] = None

class DeleteReq(BaseModel):
    pu_item_ids: List[int]
    admin_code: str

class PUCardUpdate(BaseModel):
    status: Optional[str] = None
    faza: Optional[str] = None
    voltage: Optional[str] = None
    power: Optional[float] = None
    contract_number: Optional[str] = None
    contract_date: Optional[date] = None
    plan_date: Optional[date] = None
    consumer: Optional[str] = None
    address: Optional[str] = None
    ls_number: Optional[str] = None
    smr_executor: Optional[str] = None
    smr_date: Optional[date] = None
    smr_master_id: Optional[int] = None
    ttr_ou_id: Optional[int] = None
    ttr_ol_id: Optional[int] = None
    ttr_or_id: Optional[int] = None
    ttr_tt_id: Optional[int] = None
    ttr_esk_id: Optional[int] = None
    trubostoyka: Optional[bool] = None
    form_factor: Optional[str] = None
    va_type: Optional[str] = None
    lsr_number: Optional[str] = None
    price_no_nds: Optional[float] = None
    price_with_nds: Optional[float] = None
    lsr_truba: Optional[str] = None
    price_truba_no_nds: Optional[float] = None
    price_truba_with_nds: Optional[float] = None
    lsr_va: Optional[str] = None
    price_va_no_nds: Optional[float] = None
    price_va_with_nds: Optional[float] = None
    request_number: Optional[str] = None
    request_contract: Optional[str] = None
    work_type_name: Optional[str] = None
    has_va: Optional[bool] = None
    va_nominal_id: Optional[int] = None
    va_quantity: Optional[int] = None
    has_tt: Optional[bool] = None
    tt_nominal_id: Optional[int] = None

# ==================== ПРИЛОЖЕНИЕ ====================
app = FastAPI(title="Система учета ПУ")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_credentials=True, allow_methods=["*"], allow_headers=["*"])

# Создание/миграция схемы выполняется в ensure_db_schema() при старте (см. конец файла)

# ==================== API: AUTH ====================

@app.get("/api/pu/check-contract")
def check_contract_duplicate(
    contract_number: str,
    exclude_id: Optional[int] = None,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user)
):
    """Проверка дубликата номера договора"""
    if not contract_number or len(contract_number) < 10:
        return {"duplicate": False}
    
    q = db.query(PUItem).filter(PUItem.contract_number == contract_number)
    if exclude_id:
        q = q.filter(PUItem.id != exclude_id)
    
    existing = q.first()
    if existing:
        return {
            "duplicate": True,
            "existing_serial": existing.serial_number,
            "existing_unit": existing.current_unit.name if existing.current_unit else None
        }
    return {"duplicate": False}

@app.get("/")
def root():
    return {"status": "ok", "message": "Система учета ПУ v2"}

@app.post("/api/auth/login", response_model=TokenResp)
def login(req: LoginReq, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.username == req.username).first()
    if not user or not verify_password(req.password, user.password_hash):
        raise HTTPException(401, "Неверный логин или пароль")
    return {"access_token": create_token(user.id)}

@app.get("/api/auth/me", response_model=UserResp)
def get_me(user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    return UserResp(
        id=user.id, username=user.username, full_name=user.full_name,
        role_code=user.role.code.value, role_name=user.role.name,
        unit_id=user.unit_id, 
        unit_name=user.unit.name if user.unit else None,
        unit_type=user.unit.unit_type.value if user.unit else None,
        visible_units=get_visible_units(user, db)
    )

# ==================== API: СПРАВОЧНИКИ ====================
@app.get("/api/units", response_model=List[UnitResp])
def get_units(db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    units = db.query(Unit).filter(Unit.is_active == True).all()
    return [UnitResp(id=u.id, name=u.name, code=u.code, unit_type=u.unit_type.value, short_code=u.short_code) for u in units]

@app.get("/api/roles")
def get_roles(db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    return [{"id": r.id, "name": r.name, "code": r.code.value} for r in db.query(Role).all()]

@app.get("/api/ttr/res")
def get_ttr_res(db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    """Справочник ТТР для РЭС"""
    items = db.query(TTR_RES).filter(TTR_RES.is_active == True).all()
    return [{"id": t.id, "code": t.code, "name": t.name, "ttr_type": t.ttr_type, "use_tt": t.use_tt} for t in items]

@app.get("/api/ttr/esk")
def get_ttr_esk(ttr_type: Optional[str] = None, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    """Справочник ТТР для ЭСК"""
    q = db.query(TTR_ESK).filter(TTR_ESK.is_active == True)
    if ttr_type:
        q = q.filter(TTR_ESK.ttr_type == ttr_type)
    items = q.all()
    return [{
        "id": t.id,
        "ttr_type": t.ttr_type,
        "work_type_name": t.work_type_name,
        "pu_pattern": t.pu_pattern,
        "faza": t.faza, 
        "form_factor": t.form_factor, 
        "va_type": t.va_type,
        "lsr_number": t.lsr_number,
        "price_no_nds": t.price_no_nds, 
        "price_with_nds": t.price_with_nds
    } for t in items]

@app.get("/api/ttr/esk/lookup")
def lookup_ttr_esk(
    faza: Optional[str] = None,
    form_factor: Optional[str] = None,
    va_type: Optional[str] = None,
    pu_type: Optional[str] = None,
    need_trubostoyka: bool = False,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user)
):
    """Подбор ТТР ЭСК: возвращает отдельно трубостойку и ВА"""
    
    result = {
        "trubostoyka": None,
        "va": None,
        "total_no_nds": 0,
        "total_with_nds": 0
    }
    
    # 1. Трубостойка (если нужна)
    if need_trubostoyka:
        ttr_truba = db.query(TTR_ESK).filter(
            TTR_ESK.ttr_type == "TRUBOSTOYKA",
            TTR_ESK.is_active == True
        ).first()
        if ttr_truba:
            result["trubostoyka"] = {
                "id": ttr_truba.id,
                "lsr_number": ttr_truba.lsr_number,
                "work_type_name": ttr_truba.work_type_name,
                "price_no_nds": ttr_truba.price_no_nds or 0,
                "price_with_nds": ttr_truba.price_with_nds or 0
            }
            result["total_no_nds"] += ttr_truba.price_no_nds or 0
            result["total_with_nds"] += ttr_truba.price_with_nds or 0
    
    # 2. ВА по критериям (паттерн ПУ, фаза, форм-фактор, тип ВА)
    if faza and form_factor and va_type:
        q = db.query(TTR_ESK).filter(
            TTR_ESK.ttr_type == "PU",
            TTR_ESK.faza == faza,
            TTR_ESK.form_factor == form_factor,
            TTR_ESK.va_type == va_type,
            TTR_ESK.is_active == True
        )
        
        ttr_va = None
        
        # Если передан тип ПУ, ищем по паттерну
        if pu_type:
            pu_type_upper = pu_type.upper()
            all_ttr = q.all()
            for t in all_ttr:
                if t.pu_pattern and t.pu_pattern.upper() in pu_type_upper:
                    ttr_va = t
                    break
            if not ttr_va and all_ttr:
                ttr_va = all_ttr[0]  # fallback
        else:
            ttr_va = q.first()
        
        if ttr_va:
            result["va"] = {
                "id": ttr_va.id,
                "lsr_number": ttr_va.lsr_number,
                "work_type_name": ttr_va.work_type_name,
                "price_no_nds": ttr_va.price_no_nds or 0,
                "price_with_nds": ttr_va.price_with_nds or 0
            }
            result["total_no_nds"] += ttr_va.price_no_nds or 0
            result["total_with_nds"] += ttr_va.price_with_nds or 0
    
    return result

@app.get("/api/masters")
def get_masters(unit_id: Optional[int] = None, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    """Справочник мастеров ЭСК"""
    q = db.query(ESKMaster).filter(ESKMaster.is_active == True)
    if unit_id:
        q = q.filter(ESKMaster.unit_id == unit_id)
    return [{"id": m.id, "full_name": m.full_name, "unit_id": m.unit_id, "unit_name": m.unit.name if m.unit else None} for m in q.all()]

# ==================== API: ПОЛЬЗОВАТЕЛИ (только СУЭ) ====================
@app.get("/api/users")
def get_users(db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    if not is_sue_admin(user):
        raise HTTPException(403, "Нет доступа")
    users = db.query(User).all()
    return [{
        "id": u.id, "username": u.username, "full_name": u.full_name, "is_active": u.is_active,
        "role": {"id": u.role.id, "name": u.role.name, "code": u.role.code.value} if u.role else None,
        "unit": {"id": u.unit.id, "name": u.unit.name, "unit_type": u.unit.unit_type.value} if u.unit else None
    } for u in users]

@app.post("/api/users")
def create_user(data: dict, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    if not is_sue_admin(user):
        raise HTTPException(403, "Нет доступа")
    if db.query(User).filter(User.username == data["username"]).first():
        raise HTTPException(400, "Логин уже занят")
    new_user = User(
        username=data["username"], password_hash=hash_password(data["password"]),
        full_name=data["full_name"], role_id=data["role_id"], unit_id=data.get("unit_id")
    )
    db.add(new_user)
    db.commit()
    return {"id": new_user.id}

@app.put("/api/users/{user_id}")
def update_user(user_id: int, data: dict, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    if not is_sue_admin(user):
        raise HTTPException(403, "Нет доступа")
    u = db.query(User).filter(User.id == user_id).first()
    if not u:
        raise HTTPException(404, "Не найден")
    for k, v in data.items():
        if k != "password" and hasattr(u, k):
            setattr(u, k, v)
    db.commit()
    return {"ok": True}

# ==================== API: ПУ ====================
@app.get("/api/pu/dashboard")
def dashboard(db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    try:
        visible = get_visible_units(user, db)
        
        # Получаем ID РЭС и ЭСК
        res_unit_ids = [u.id for u in db.query(Unit).filter(Unit.unit_type == UnitType.RES).all()]
        esk_unit_ids = [u.id for u in db.query(Unit).filter(Unit.unit_type.in_([UnitType.ESK, UnitType.ESK_UNIT])).all()]
        oks_unit_ids = [u.id for u in db.query(Unit).filter(Unit.unit_type.in_([UnitType.OKS, UnitType.OKS_UNIT])).all()]
        
        def get_stats(unit_ids=None):
            # Один запрос с группировкой по статусу вместо пяти отдельных COUNT
            q = db.query(PUItem.status, func.count(PUItem.id))
            if is_lab_user(user):
                regs = db.query(PURegister.id).filter(PURegister.uploaded_by == user.id)
                q = q.filter(PUItem.register_id.in_(regs))
            elif not is_sue_admin(user):
                q = q.filter(PUItem.current_unit_id.in_(visible))
            
            if unit_ids:
                q = q.filter(PUItem.current_unit_id.in_(unit_ids))
            
            counts = {}
            for status_val, cnt in q.group_by(PUItem.status).all():
                key = status_val.value if hasattr(status_val, 'value') else status_val
                counts[key] = cnt
            
            sklad = counts.get('SKLAD', 0)
            techpris = counts.get('TECHPRIS', 0)
            zamena = counts.get('ZAMENA', 0)
            izhc = counts.get('IZHC', 0)
            installed = techpris + zamena + izhc
            total = sum(counts.values())
            
            return {
                "total": total,
                "installed": installed,
                "sklad": sklad,
                "techpris": techpris,
                "zamena": zamena,
                "izhc": izhc
            }
        
        # Статистика по всем
        stats_all = get_stats()
        
        # Статистика по РЭС
        stats_res = get_stats(res_unit_ids)
        
        # Статистика по ЭСК
        stats_esk = get_stats(esk_unit_ids)
        
        # Статистика по ОКС
        stats_oks = get_stats(oks_unit_ids)
        
        # Последние загрузки
        reg_q = db.query(PURegister)
        if is_lab_user(user):
            reg_q = reg_q.filter(PURegister.uploaded_by == user.id)
        recent = reg_q.order_by(PURegister.uploaded_at.desc()).limit(5).all()
        
        # Количество на согласовании
        pending_approval = 0
        if is_res_user(user) and user.unit_id:
            res_code = user.unit.code if user.unit else ""
            if res_code:
                base = res_code.replace("RES_", "")
                pending_codes = [f"ESK_{base}", f"OKS_{base}"]
                pending_unit_ids = [u.id for u in db.query(Unit).filter(Unit.code.in_(pending_codes)).all()]
                if pending_unit_ids:
                    pending_approval = db.query(PUItem).filter(
                        PUItem.current_unit_id.in_(pending_unit_ids),
                        PUItem.approval_status == ApprovalStatus.PENDING
                    ).count()
        
        return {
            "all": stats_all,
            "res": stats_res,
            "esk": stats_esk,
            "oks": stats_oks,
            "pending_approval": pending_approval,
            "recent_registers": [{"id": r.id, "filename": r.filename, "items_count": r.items_count, "uploaded_at": r.uploaded_at} for r in recent]
        }
        
    except Exception as e:
        print(f"Dashboard error: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(500, f"Ошибка: {str(e)}")

@app.get("/api/pu/analysis")
def get_analysis(
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user)
):
    """Анализ остатков по подразделениям с разбивкой по назначению/форм-фактору/фазности"""
    try:
        from datetime import datetime
        
        start_date = None
        end_date = None
        if date_from:
            start_date = datetime.strptime(date_from, '%Y-%m-%d')
        if date_to:
            end_date = datetime.strptime(date_to, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
        
        # Категории для разбивки
        naznachenie_list = ['IZHC', 'TECHPRIS', 'ZAMENA']
        form_factors = ['split', 'classic']
        faza_split = ['1ф', '3ф']
        faza_classic = ['1ф', '3ф', '3фтт']
        
        def get_unit_stats(unit_id):
            base_filters = [PUItem.current_unit_id == unit_id]
            if start_date:
                base_filters.append(PUItem.created_at >= start_date)
            if end_date:
                base_filters.append(PUItem.created_at <= end_date)
            
            actioned_cond = or_(
                (PUItem.tz_number != None) & (PUItem.tz_number != ""),
                (PUItem.request_number != None) & (PUItem.request_number != "")
            )
            
            q = db.query(PUItem).filter(*base_filters)
            total = q.count()
            sklad = q.filter(PUItem.status == PUStatus.SKLAD).count()
            installed = q.filter(PUItem.status != PUStatus.SKLAD).count()
            actioned = q.filter(actioned_cond).count()
            
            # Группирующие запросы вместо десятков отдельных COUNT.
            # status_map: (статус, форм-фактор, фаза) -> кол-во (для установленных/актированных)
            # sklad_map:  (назначение, форм-фактор, фаза) -> кол-во (для склада)
            def to_str(v):
                return v.value if hasattr(v, 'value') else v
            
            status_rows = db.query(
                PUItem.status, PUItem.form_factor, PUItem.faza, func.count(PUItem.id)
            ).filter(*base_filters).group_by(
                PUItem.status, PUItem.form_factor, PUItem.faza
            ).all()
            status_map = {(to_str(s), ff, fz): c for s, ff, fz, c in status_rows}
            
            actioned_rows = db.query(
                PUItem.status, PUItem.form_factor, PUItem.faza, func.count(PUItem.id)
            ).filter(*base_filters, actioned_cond).group_by(
                PUItem.status, PUItem.form_factor, PUItem.faza
            ).all()
            actioned_map = {(to_str(s), ff, fz): c for s, ff, fz, c in actioned_rows}
            
            sklad_rows = db.query(
                PUItem.naznachenie, PUItem.form_factor, PUItem.faza, func.count(PUItem.id)
            ).filter(*base_filters, PUItem.status == PUStatus.SKLAD).group_by(
                PUItem.naznachenie, PUItem.form_factor, PUItem.faza
            ).all()
            sklad_map = {(naz, ff, fz): c for naz, ff, fz, c in sklad_rows}
            
            # Детальная разбивка (та же логика и те же цифры, что и раньше):
            # total: склад → по назначению + не-склад → по статусу
            # installed: по статусу (!= склад); actioned: по статусу + фильтр актирования
            # sklad: по назначению при статусе склад
            breakdown = {}
            for section_key in ['total', 'installed', 'actioned', 'sklad']:
                section = {}
                for naz in naznachenie_list:
                    naz_data = {}
                    for ff in form_factors:
                        fazas = faza_split if ff == 'split' else faza_classic
                        ff_data = {}
                        for fz in fazas:
                            if section_key == 'total':
                                ff_data[fz] = sklad_map.get((naz, ff, fz), 0) + status_map.get((naz, ff, fz), 0)
                            elif section_key == 'installed':
                                ff_data[fz] = status_map.get((naz, ff, fz), 0)
                            elif section_key == 'actioned':
                                ff_data[fz] = actioned_map.get((naz, ff, fz), 0)
                            elif section_key == 'sklad':
                                ff_data[fz] = sklad_map.get((naz, ff, fz), 0)
                        naz_data[ff] = ff_data
                    section[naz] = naz_data
                breakdown[section_key] = section
            
            return {
                "total": total,
                "installed": installed,
                "actioned": actioned,
                "sklad": sklad,
                "breakdown": breakdown
            }
        
        result = {"res": [], "esk": [], "oks": []}
        
        if is_res_user(user) or is_esk_user(user) or is_oks_user(user):
            if user.unit_id:
                unit = db.query(Unit).filter(Unit.id == user.unit_id).first()
                if unit:
                    stats = get_unit_stats(unit.id)
                    item = {"id": unit.id, "name": unit.name, **stats}
                    if unit.unit_type == UnitType.RES:
                        result["res"].append(item)
                    elif unit.unit_type in (UnitType.OKS, UnitType.OKS_UNIT):
                        result["oks"].append(item)
                    else:
                        result["esk"].append(item)
            return result
        
        res_units = db.query(Unit).filter(Unit.unit_type == UnitType.RES).order_by(Unit.name).all()
        esk_units = db.query(Unit).filter(Unit.unit_type.in_([UnitType.ESK, UnitType.ESK_UNIT])).order_by(Unit.name).all()
        oks_units = db.query(Unit).filter(Unit.unit_type.in_([UnitType.OKS, UnitType.OKS_UNIT])).order_by(Unit.name).all()
        
        def empty_breakdown():
            bd = {}
            for sk in ['total', 'installed', 'actioned', 'sklad']:
                section = {}
                for naz in naznachenie_list:
                    naz_data = {}
                    for ff in form_factors:
                        fazas = faza_split if ff == 'split' else faza_classic
                        naz_data[ff] = {fz: 0 for fz in fazas}
                    section[naz] = naz_data
                bd[sk] = section
            return bd
        
        def sum_breakdowns(total_bd, add_bd):
            for sk in total_bd:
                for naz in total_bd[sk]:
                    for ff in total_bd[sk][naz]:
                        for fz in total_bd[sk][naz][ff]:
                            total_bd[sk][naz][ff][fz] += add_bd.get(sk, {}).get(naz, {}).get(ff, {}).get(fz, 0)
        
        res_total = {"total": 0, "installed": 0, "actioned": 0, "sklad": 0, "breakdown": empty_breakdown()}
        for unit in res_units:
            stats = get_unit_stats(unit.id)
            result["res"].append({"id": unit.id, "name": unit.name, **stats})
            res_total["total"] += stats["total"]
            res_total["installed"] += stats["installed"]
            res_total["actioned"] += stats["actioned"]
            res_total["sklad"] += stats["sklad"]
            sum_breakdowns(res_total["breakdown"], stats["breakdown"])
        
        esk_total = {"total": 0, "installed": 0, "actioned": 0, "sklad": 0, "breakdown": empty_breakdown()}
        for unit in esk_units:
            stats = get_unit_stats(unit.id)
            result["esk"].append({"id": unit.id, "name": unit.name, **stats})
            esk_total["total"] += stats["total"]
            esk_total["installed"] += stats["installed"]
            esk_total["actioned"] += stats["actioned"]
            esk_total["sklad"] += stats["sklad"]
            sum_breakdowns(esk_total["breakdown"], stats["breakdown"])
        
        oks_total = {"total": 0, "installed": 0, "actioned": 0, "sklad": 0, "breakdown": empty_breakdown()}
        for unit in oks_units:
            stats = get_unit_stats(unit.id)
            result["oks"].append({"id": unit.id, "name": unit.name, **stats})
            oks_total["total"] += stats["total"]
            oks_total["installed"] += stats["installed"]
            oks_total["actioned"] += stats["actioned"]
            oks_total["sklad"] += stats["sklad"]
            sum_breakdowns(oks_total["breakdown"], stats["breakdown"])
        
        grand_bd = empty_breakdown()
        sum_breakdowns(grand_bd, res_total["breakdown"])
        sum_breakdowns(grand_bd, esk_total["breakdown"])
        sum_breakdowns(grand_bd, oks_total["breakdown"])
        
        result["res_total"] = res_total
        result["esk_total"] = esk_total
        result["oks_total"] = oks_total
        result["grand_total"] = {
            "total": res_total["total"] + esk_total["total"] + oks_total["total"],
            "installed": res_total["installed"] + esk_total["installed"] + oks_total["installed"],
            "actioned": res_total["actioned"] + esk_total["actioned"] + oks_total["actioned"],
            "sklad": res_total["sklad"] + esk_total["sklad"] + oks_total["sklad"],
            "breakdown": grand_bd
        }
        
        return result
        
    except Exception as e:
        print(f"Analysis error: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(500, f"Ошибка: {str(e)}")

@app.get("/api/pu/items")
def get_items(
    page: int = 1, size: int = 50,
    search: Optional[str] = None, 
    status: Optional[str] = None, 
    unit_id: Optional[int] = None,
    unit_type_filter: Optional[str] = None,
    contract: Optional[str] = None,
    ls: Optional[str] = None,
    filter: Optional[str] = None,
    sort_field: Optional[str] = None,
    sort_dir: Optional[str] = 'desc',
    db: Session = Depends(get_db), 
    user: User = Depends(get_current_user)
):
    visible = get_visible_units(user, db)
    q = db.query(PUItem).options(joinedload(PUItem.current_unit))
    
    if is_lab_user(user):
        regs = db.query(PURegister.id).filter(PURegister.uploaded_by == user.id)
        q = q.filter(PUItem.register_id.in_(regs))
    elif not is_sue_admin(user):
        q = q.filter(PUItem.current_unit_id.in_(visible))
    
    if search:
        q = q.filter(PUItem.serial_number.ilike(f"%{search}%"))
    if status:
        q = q.filter(PUItem.status == status)
    if unit_id:
        q = q.filter(PUItem.current_unit_id == unit_id)
    # Фильтр по типу подразделения
    if unit_type_filter == 'res':
        res_units = db.query(Unit.id).filter(Unit.unit_type == UnitType.RES)
        q = q.filter(PUItem.current_unit_id.in_(res_units))
    elif unit_type_filter == 'esk':
        esk_units = db.query(Unit.id).filter(Unit.unit_type.in_([UnitType.ESK, UnitType.ESK_UNIT]))
        q = q.filter(PUItem.current_unit_id.in_(esk_units))
    elif unit_type_filter == 'oks':
        oks_units = db.query(Unit.id).filter(Unit.unit_type.in_([UnitType.OKS, UnitType.OKS_UNIT]))
        q = q.filter(PUItem.current_unit_id.in_(oks_units))
    if contract:
        q = q.filter(PUItem.contract_number.ilike(f"%{contract}%"))
    if ls:
        q = q.filter(PUItem.ls_number.ilike(f"%{ls}%"))

    # Фильтр по типу реестра
    # Фильтр по типу реестра
    if filter == 'sklad':
        # Только склад
        q = q.filter(PUItem.status == PUStatus.SKLAD)
    elif filter == 'done':
        # Завершённые СМР — любой статус кроме склада
        q = q.filter(PUItem.status != PUStatus.SKLAD)
    elif filter == 'actioned':
        # Актированные — есть ТЗ или Заявка
        from sqlalchemy import or_
        q = q.filter(
            or_(
                (PUItem.tz_number != None) & (PUItem.tz_number != ""),
                (PUItem.request_number != None) & (PUItem.request_number != "")
            )
        )

    total = q.count()
    
    # Сортировка
    sort_mapping = {
        'serial_number': PUItem.serial_number,
        'pu_type': PUItem.pu_type,
        'status': PUItem.status,
        'tz_number': PUItem.tz_number,
        'request_number': PUItem.request_number,
        'approval_status': PUItem.approval_status,
        'created_at': PUItem.created_at,
    }
    
    sort_column = sort_mapping.get(sort_field, PUItem.created_at)
    if sort_dir == 'asc':
        q = q.order_by(sort_column.asc())
    else:
        q = q.order_by(sort_column.desc())
    
    items = q.offset((page-1)*size).limit(size).all()
    
    return {
    "items": [{
        "id": i.id, "serial_number": i.serial_number, "pu_type": i.pu_type,
        "status": i.status.value, "naznachenie": i.naznachenie,
        "current_unit_id": i.current_unit_id,
        "current_unit_name": i.current_unit.name if i.current_unit else None,
        "current_unit_type": i.current_unit.unit_type.value if i.current_unit else None,
        "tz_number": i.tz_number, "request_number": i.request_number,
        "contract_number": i.contract_number,
        "ls_number": i.ls_number, "consumer": i.consumer,
        "smr_date": i.smr_date.isoformat() if i.smr_date else None,
        "approval_status": i.approval_status.value if i.approval_status else None,
        "uploaded_at": i.register.uploaded_at if i.register else None
    } for i in items],
    "total": total, "page": page, "size": size, "pages": (total + size - 1) // size
}

@app.get("/api/pu/export")
def export_pu_items(
    search: Optional[str] = None,
    status: Optional[str] = None,
    unit_id: Optional[int] = None,
    unit_type_filter: Optional[str] = None,
    contract: Optional[str] = None,
    ls: Optional[str] = None,
    filter: Optional[str] = None,  # all, work, done
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user)
):
    """Выгрузка реестра ПУ в Excel"""
    try:
        visible = get_visible_units(user, db)
        q = db.query(PUItem).options(joinedload(PUItem.current_unit), joinedload(PUItem.ttr_esk))
        
        # Те же фильтры что и в get_items
        if is_lab_user(user):
            regs = db.query(PURegister.id).filter(PURegister.uploaded_by == user.id)
            q = q.filter(PUItem.register_id.in_(regs))
        elif not is_sue_admin(user):
            q = q.filter(PUItem.current_unit_id.in_(visible))
        
        if search:
            q = q.filter(PUItem.serial_number.ilike(f"%{search}%"))
        if status:
            q = q.filter(PUItem.status == status)
        if unit_id:
            q = q.filter(PUItem.current_unit_id == unit_id)
        if unit_type_filter == 'res':
            res_units = db.query(Unit.id).filter(Unit.unit_type == UnitType.RES)
            q = q.filter(PUItem.current_unit_id.in_(res_units))
        elif unit_type_filter == 'esk':
            esk_units = db.query(Unit.id).filter(Unit.unit_type.in_([UnitType.ESK, UnitType.ESK_UNIT]))
            q = q.filter(PUItem.current_unit_id.in_(esk_units))
        if contract:
            q = q.filter(PUItem.contract_number.ilike(f"%{contract}%"))
        if ls:
            q = q.filter(PUItem.ls_number.ilike(f"%{ls}%"))
        
        if filter == 'sklad':
            q = q.filter(PUItem.status == PUStatus.SKLAD)
        elif filter == 'done':
            q = q.filter(PUItem.status != PUStatus.SKLAD)
        elif filter == 'actioned':
            from sqlalchemy import or_
            q = q.filter(
                or_(
                    (PUItem.tz_number != None) & (PUItem.tz_number != ""),
                    (PUItem.request_number != None) & (PUItem.request_number != "")
                )
            )
        elif filter == 'done':
            from sqlalchemy import or_
            q = q.filter(
                or_(
                    (PUItem.tz_number != None) & (PUItem.tz_number != ""),
                    (PUItem.request_number != None) & (PUItem.request_number != ""),
                    PUItem.approval_status == ApprovalStatus.APPROVED
                )
            )
        
        items = q.order_by(PUItem.created_at.desc()).all()
        
        # Создаём Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Реестр ПУ"
        
        # Стили
        header_font = Font(bold=True, color="FFFFFF", size=10)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # Заголовки
        headers = [
            ("№", 5),
            ("Серийный номер", 20),
            ("Тип ПУ", 40),
            ("Подразделение", 20),
            ("Статус", 12),
            ("Назначение", 12),
            ("Фазность", 10),
            ("Напряжение", 12),
            ("Мощность", 10),
            ("№ Договора", 22),
            ("Потребитель", 25),
            ("Адрес", 35),
            ("ЛС", 15),
            ("Трубостойка", 12),
            ("№ ТЗ", 15),
            ("№ Заявки", 12),
            ("Согласование", 15),
            ("Вид работ (ЛСР)", 30),
            ("Дата СМР", 12),
            ("Дата загрузки", 12),
        ]
        
        for col, (header, width) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col)].width = width
        
        ws.row_dimensions[1].height = 35
        
        # Данные
        status_labels = {
            'SKLAD': 'Склад', 'TECHPRIS': 'Техприс', 
            'ZAMENA': 'Замена', 'IZHC': 'ИЖЦ', 'INSTALLED': 'Установлен'
        }
        approval_labels = {
            'APPROVED': 'Согласовано', 'PENDING': 'На согласовании', 
            'REJECTED': 'Отклонено', 'NONE': '—'
        }
        
        for idx, item in enumerate(items, 1):
            row = idx + 1
            naznachenie_labels = {'IZHC': 'ИЖЦ', 'TECHPRIS': 'Техприс', 'ZAMENA': 'Замена'}
            data = [
                idx,
                item.serial_number or "",
                item.pu_type or "",
                item.current_unit.name if item.current_unit else "",
                status_labels.get(item.status.value, item.status.value) if item.status else "",
                naznachenie_labels.get(item.naznachenie, item.naznachenie or ""),
                item.faza or "",
                item.voltage or "",
                item.power or "",
                item.contract_number or "",
                item.consumer or "",
                item.address or "",
                item.ls_number or "",
                "Да" if item.trubostoyka else "Нет", 
                item.tz_number or "",
                item.request_number or "",
                approval_labels.get(item.approval_status.value if item.approval_status else 'NONE', '—'),
                item.work_type_name or (item.ttr_esk.work_type_name if item.ttr_esk else "") or "",
                item.smr_date.strftime("%d.%m.%Y") if item.smr_date else "",
                item.created_at.strftime("%d.%m.%Y") if item.created_at else "",
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center", wrap_text=True)
            
            ws.row_dimensions[row].height = 25
        
        # Сохраняем
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Имя файла (ASCII для совместимости + UTF-8 для красоты)
        filter_name_ascii = {"sklad": "Sklad", "done": "Zavershennye_SMR", "actioned": "Aktirovannye"}.get(filter, "Vse")
        filter_name_rus = {"sklad": "Склад", "done": "Завершенные_СМР", "actioned": "Актированные"}.get(filter, "Все")

        filename_ascii = f"Reestr_PU_{filter_name_ascii}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        filename_rus = f"Реестр_ПУ_{filter_name_rus}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename=\"{filename_ascii}\"; filename*=UTF-8''{quote(filename_rus)}"
            }
        )
        
    except Exception as e:
        print(f"Export error: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(500, f"Ошибка экспорта: {str(e)}")

@app.get("/api/pu/detect-type")
def detect_type(pu_type: str, db: Session = Depends(get_db)):
    """Определить фазность и напряжение по типу ПУ"""
    result = detect_pu_type_params(pu_type, db)
    return result

@app.get("/api/pu/debug-detect/{serial}")
def debug_detect(serial: str, db: Session = Depends(get_db)):
    """Отладка: почему ПУ не находится по паттерну"""
    item = db.query(PUItem).filter(PUItem.serial_number == serial).first()
    if not item:
        return {"error": f"ПУ {serial} не найден"}
    
    pu_norm = normalize_pu_string(item.pu_type)
    
    patterns = db.query(PUTypeReference).filter(PUTypeReference.is_active == True).all()
    
    matches = []
    for p in patterns:
        if not p.pattern:
            continue
        p_norm = normalize_pu_string(p.pattern)
        
        # Проверяем разные типы совпадений
        exact_in = p_norm in pu_norm
        starts = pu_norm.startswith(p_norm)
        
        # Токены
        pu_tokens = set(pu_norm.split())
        p_tokens = set(p_norm.split())
        common = p_tokens & pu_tokens
        score = len(common) / len(p_tokens) if p_tokens else 0
        
        matches.append({
            "pattern_raw": p.pattern,
            "pattern_norm": p_norm,
            "exact_in": exact_in,
            "starts_with": starts,
            "token_score": round(score, 2),
            "common_tokens": list(common),
            "faza": p.faza,
            "form_factor": p.form_factor,
        })
    
    # Сортируем по score
    matches.sort(key=lambda x: x["token_score"], reverse=True)
    
    return {
        "serial": serial,
        "pu_type_raw": item.pu_type,
        "pu_type_norm": pu_norm,
        "current_faza": item.faza,
        "current_ff": item.form_factor,
        "detected": detect_pu_type_params(item.pu_type, db),
        "top_matches": matches[:5]
    }


@app.get("/api/pu/items/{item_id}")
def get_item_detail(item_id: int, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    """Полная карточка ПУ"""
    item = db.query(PUItem).filter(PUItem.id == item_id).first()
    if not item:
        raise HTTPException(404, "ПУ не найден")
    
    return {
        "id": item.id,
        "serial_number": item.serial_number,
        "pu_type": item.pu_type,
        "status": item.status.value,
        "naznachenie": item.naznachenie,
        "current_unit_id": item.current_unit_id,
        "current_unit_name": item.current_unit.name if item.current_unit else None,
        "current_unit_type": item.current_unit.unit_type.value if item.current_unit else None,
        "tz_number": item.tz_number,
        "faza": item.faza,
        "voltage": item.voltage,
        "power": item.power,
        "contract_number": item.contract_number,
        "contract_date": item.contract_date.isoformat() if item.contract_date else None,
        "plan_date": item.plan_date.isoformat() if item.plan_date else None,
        "consumer": item.consumer,
        "address": item.address,
        "ls_number": item.ls_number,
        "smr_executor": item.smr_executor,
        "smr_date": item.smr_date.isoformat() if item.smr_date else None,
        "smr_master_id": item.smr_master_id,
        "ttr_ou_id": item.ttr_ou_id,
        "ttr_ol_id": item.ttr_ol_id,
        "ttr_or_id": item.ttr_or_id,
        "ttr_tt_id": item.ttr_tt_id,
        "ttr_esk_id": item.ttr_esk_id,
        "trubostoyka": item.trubostoyka,
        "materials_used": item.materials_used,
        "approval_status": item.approval_status.value if item.approval_status else None,
        "request_number": item.request_number,
        "ttr_esk_id": item.ttr_esk_id,
        "trubostoyka": item.trubostoyka,
        "form_factor": item.form_factor,
        "va_type": item.va_type,
        "lsr_number": item.lsr_number,
        "price_no_nds": item.price_no_nds,
        "price_with_nds": item.price_with_nds,
        "lsr_truba": item.lsr_truba,
        "price_truba_no_nds": item.price_truba_no_nds,
        "price_truba_with_nds": item.price_truba_with_nds,
        "lsr_va": item.lsr_va,
        "price_va_no_nds": item.price_va_no_nds,
        "price_va_with_nds": item.price_va_with_nds,
        "request_contract": item.request_contract,
        "work_type_name": item.work_type_name,
        "rejection_comment": item.rejection_comment,
        "has_va": item.has_va,
        "va_nominal_id": item.va_nominal_id,
        "va_nominal_name": item.va_nominal.name if item.va_nominal else None,
        "has_tt": item.has_tt,
        "tt_nominal_id": item.tt_nominal_id,
        "tt_nominal_name": item.tt_nominal.name if item.tt_nominal else None,
    }


@app.get("/api/pu/items/{item_id}/review")
def get_item_review(item_id: int, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    """Детали ПУ для проверки на согласовании (РЭС/СУЭ): ТТР, материалы, оборудование"""
    if not is_res_user(user) and not is_sue_admin(user):
        raise HTTPException(403, "Нет доступа")
    
    item = db.query(PUItem).filter(PUItem.id == item_id).first()
    if not item:
        raise HTTPException(404, "ПУ не найден")
    
    def ttr_label(ttr_id):
        if not ttr_id:
            return None
        t = db.query(TTR_RES).filter(TTR_RES.id == ttr_id).first()
        if not t:
            return None
        return f"{t.code} — {t.name}" if t.name else t.code
    
    # Фактические материалы
    materials = []
    for pm in db.query(PUMaterial).filter(PUMaterial.pu_item_id == item_id).all():
        m = db.query(Material).filter(Material.id == pm.material_id).first()
        materials.append({
            "name": m.name if m else "—",
            "unit": m.unit if m else "",
            "quantity": pm.quantity,
            "used": pm.used,
        })
    
    master_name = None
    if item.smr_master_id:
        mm = db.query(ESKMaster).filter(ESKMaster.id == item.smr_master_id).first()
        master_name = mm.full_name if mm else None
    
    return {
        "status": item.status.value if item.status else None,
        "naznachenie": item.naznachenie,
        "faza": item.faza,
        "voltage": item.voltage,
        "power": item.power,
        "contract_number": item.contract_number,
        "contract_date": item.contract_date.isoformat() if item.contract_date else None,
        "plan_date": item.plan_date.isoformat() if item.plan_date else None,
        "consumer": item.consumer,
        "address": item.address,
        "ls_number": item.ls_number,
        "smr_executor": item.smr_executor,
        "smr_date": item.smr_date.isoformat() if item.smr_date else None,
        "smr_master": master_name,
        "trubostoyka": item.trubostoyka,
        # ТТР РЭС (коды)
        "ttr_ou": ttr_label(item.ttr_ou_id),
        "ttr_ol": ttr_label(item.ttr_ol_id),
        "ttr_or": ttr_label(item.ttr_or_id),
        "ttr_tt": ttr_label(item.ttr_tt_id),
        # Оборудование
        "has_va": item.has_va,
        "va_nominal": item.va_nominal.name if item.va_nominal else None,
        "va_quantity": item.va_quantity,
        "has_tt": item.has_tt,
        "tt_nominal": item.tt_nominal.name if item.tt_nominal else None,
        # Материалы
        "materials": materials,
        # Поля ЭСК (если раскрывают карточку ЭСК)
        "form_factor": item.form_factor,
        "va_type": item.va_type,
        "lsr_va": item.lsr_va,
        "lsr_truba": item.lsr_truba,
    }

@app.put("/api/pu/items/{item_id}")
def update_item(item_id: int, data: PUCardUpdate, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    """Обновление карточки ПУ"""
    item = db.query(PUItem).filter(PUItem.id == item_id).first()
    if not item:
        raise HTTPException(404, "ПУ не найден")
    
    # Проверка доступа - ЭСК Админ только просмотр
    # СУЭ Админ может редактировать карточки РЭС
    sue_admin = is_sue_admin(user)
    if sue_admin:
        # СУЭ может редактировать только карточки РЭС подразделений
        unit = db.query(Unit).filter(Unit.id == item.current_unit_id).first()
        if not unit or unit.unit_type != 'RES':
            raise HTTPException(403, "СУЭ Админ может редактировать только карточки РЭС")
    elif is_esk_admin(user):
        raise HTTPException(403, "ЭСК Админ может только просматривать и перемещать ПУ")
    
    visible = get_visible_units(user, db)
    if item.current_unit_id not in visible:
        raise HTTPException(403, "Нет доступа к этому ПУ")
    # Согласованные ПУ нельзя редактировать (кроме СУЭ админа)
    if item.approval_status == ApprovalStatus.APPROVED and not sue_admin:
        raise HTTPException(403, "Согласованные ПУ нельзя редактировать")
    
    # Валидация договора
    if data.contract_number:
        pattern = r'^\d{5}-\d{2}-\d{8}-\d$'
        if not re.match(pattern, data.contract_number):
            raise HTTPException(400, "Неверный формат договора. Ожидается: ххххх-хх-хххххххх-х")
        # Проверка дубликата
        existing = db.query(PUItem).filter(
            PUItem.contract_number == data.contract_number,
            PUItem.id != item_id
        ).first()
        if existing:
            raise HTTPException(400, f"Договор уже существует в системе (ПУ {existing.serial_number})")
    
    # Автозаполнение фазности и напряжения при смене статуса со Склада
    if data.status and data.status != 'SKLAD' and item.status == PUStatus.SKLAD:
        detected = detect_pu_type_params(item.pu_type, db)
        if detected:
            if not data.faza and not item.faza and detected.get('faza'):
                data.faza = detected['faza']
            if not data.voltage and not item.voltage and detected.get('voltage'):
                data.voltage = detected['voltage']
    
    # Запоминаем старые ТТР для очистки материалов
    old_ttr_ids = {item.ttr_ou_id, item.ttr_ol_id, item.ttr_or_id, item.ttr_tt_id}
    
    # Обновляем поля (для СУЭ админа разрешаем запись None — сброс полей)
    nullable_fields = {'ttr_ou_id', 'ttr_ol_id', 'ttr_or_id', 'ttr_tt_id', 'ttr_esk_id',
                       'smr_executor', 'smr_date', 'smr_master_id', 'contract_number', 
                       'contract_date', 'plan_date', 'consumer', 'address', 'ls_number',
                       'va_nominal_id', 'tt_nominal_id', 'power', 'form_factor', 'va_type',
                       'lsr_number', 'lsr_va', 'lsr_truba',
                       'price_no_nds', 'price_with_nds', 'price_va_no_nds', 'price_va_with_nds',
                       'price_truba_no_nds', 'price_truba_with_nds'}
    
    for key, value in data.dict(exclude_unset=True).items():
        if value is not None:
            setattr(item, key, value)
        elif key in nullable_fields:
            # Разрешаем сброс в NULL для этих полей
            setattr(item, key, None)
    
    # Если ТТР были сброшены — очищаем привязанные материалы
    new_ttr_ids = {item.ttr_ou_id, item.ttr_ol_id, item.ttr_or_id, item.ttr_tt_id}
    if old_ttr_ids != new_ttr_ids:
        # Если все ТТР сброшены — удаляем все материалы
        if not any([item.ttr_ou_id, item.ttr_ol_id, item.ttr_or_id, item.ttr_tt_id]):
            db.query(PUMaterial).filter(PUMaterial.pu_item_id == item_id).delete()
    
    db.commit()
    return {"ok": True}

@app.get("/api/pu/registers")
def get_registers(db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    q = db.query(PURegister)
    if is_lab_user(user):
        q = q.filter(PURegister.uploaded_by == user.id)
    elif not is_sue_admin(user):
        return []
    regs = q.order_by(PURegister.uploaded_at.desc()).all()
    return [{"id": r.id, "filename": r.filename, "items_count": r.items_count, "uploaded_at": r.uploaded_at} for r in regs]

@app.get("/api/pu/upload-template")
def download_upload_template():
    """Скачать шаблон Excel для загрузки реестра ПУ"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Реестр ПУ"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    example_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    note_font = Font(italic=True, color="808080", size=9)

    headers = [
        ("Заводской номер ПУ", 25),
        ("Тип прибора учета", 45),
        ("Подразделение", 25),
        ("Назначение", 18),
    ]

    for col, (header, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.row_dimensions[1].height = 30

    examples = [
        ("12345678", "Меркурий 234 ARTM-01 PB.G 5(100)А", "Восточный РЭС", "ИЖЦ"),
        ("87654321", "Нева МТ 314 1.0 AR E4BS26 5(100)А", "Западный РЭС", "Техприс"),
        ("11223344", "Меркурий 201.8TLO 5(80)А", "Южный РЭС", "Замена"),
    ]

    for row_idx, example in enumerate(examples, 2):
        for col_idx, value in enumerate(example, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = example_fill
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
        ws.row_dimensions[row_idx].height = 22

    note_row = len(examples) + 3
    ws.cell(row=note_row, column=1, value="Примечания:").font = Font(bold=True, size=9)
    ws.cell(row=note_row + 1, column=1, value="• Колонка «Заводской номер ПУ» — обязательна").font = note_font
    ws.cell(row=note_row + 2, column=1, value="• Назначение: ИЖЦ, Техприс или Замена").font = note_font
    ws.cell(row=note_row + 3, column=1, value="• Жёлтые строки — примеры, удалите их перед загрузкой").font = note_font

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename_ascii = "Shablon_Zagruzki_PU.xlsx"
    filename_rus = "Шаблон_Загрузки_ПУ.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename_ascii}"; filename*=UTF-8\'\'{quote(filename_rus)}'
        }
    )

@app.post("/api/pu/upload")
async def upload_register(file: UploadFile = File(...), db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    """Загрузка реестра ПУ - только Лаборатория"""
    if not is_lab_user(user):
        raise HTTPException(403, "Только Лаборатория может загружать реестры")
    
    contents = await file.read()
    xl = pd.ExcelFile(io.BytesIO(contents))
    
    # Ищем лист с данными
    df = None
    for sheet in xl.sheet_names:
        temp_df = pd.read_excel(xl, sheet_name=sheet)
        for col in temp_df.columns:
            if 'заводской' in str(col).lower() or ('номер' in str(col).lower() and 'пу' in str(col).lower()):
                df = temp_df
                break
        if df is not None:
            break
    if df is None:
        df = pd.read_excel(io.BytesIO(contents))
    
    register = PURegister(filename=file.filename, uploaded_by=user.id, items_count=0)
    db.add(register)
    db.commit()
    
    # Поиск колонок
    serial_col = type_col = unit_col = naznachenie_col = None
    for col in df.columns:
        col_lower = str(col).lower()
        if 'заводской' in col_lower or ('номер' in col_lower and 'пу' in col_lower):
            serial_col = col
        elif 'тип' in col_lower:
            type_col = col
        elif 'подразделение' in col_lower:
            unit_col = col
        elif 'назначение' in col_lower:
            naznachenie_col = col
    
    if not serial_col:
        raise HTTPException(400, "Не найдена колонка 'Заводской номер ПУ'")
    
    # Словарь подразделений
    units_map = {}
    for u in db.query(Unit).all():
        units_map[u.name.lower()] = u
        if u.code:
            units_map[u.code.lower()] = u
    
    count = 0
    skipped_duplicates = 0
    duplicate_serials = []
    # Предзагружаем все существующие заводские номера одним запросом
    # (вместо отдельного SELECT на каждую строку файла)
    existing_serials = set(s for (s,) in db.query(PUItem.serial_number).all())
    for _, row in df.iterrows():
        serial = str(row.get(serial_col, '')).strip()
        if not serial or serial == 'nan':
            continue
    
        # Проверка дубликата серийного номера (в БД и в пределах текущего файла)
        if serial in existing_serials:
            skipped_duplicates += 1
            duplicate_serials.append(serial)
            continue
        
        pu_type = str(row.get(type_col, '')).strip() if type_col else None
        if pu_type == 'nan':
            pu_type = None
        
        target_unit = None
        if unit_col:
            unit_name = str(row.get(unit_col, '')).strip().lower()
            if unit_name and unit_name != 'nan':
                target_unit = units_map.get(unit_name)
                if not target_unit:
                    for key, u in units_map.items():
                        if unit_name in key or key in unit_name:
                            target_unit = u
                            break
        
        # Определяем назначение
        naznachenie_val = None
        if naznachenie_col:
            naz_raw = str(row.get(naznachenie_col, '')).strip().lower()
            if naz_raw and naz_raw != 'nan':
                naz_map = {'ижц': 'IZHC', 'техприс': 'TECHPRIS', 'замена': 'ZAMENA'}
                naznachenie_val = naz_map.get(naz_raw, naz_raw.upper())

        # Автоопределение фазности, форм-фактора, напряжения по типу ПУ
        detected = detect_pu_type_params(pu_type, db) if pu_type else {}

        # По умолчанию статус СКЛАД
        item = PUItem(
            register_id=register.id,
            pu_type=pu_type[:500] if pu_type else None,
            serial_number=serial,
            target_unit_id=target_unit.id if target_unit else None,
            current_unit_id=target_unit.id if target_unit else None,
            status=PUStatus.SKLAD,
            naznachenie=naznachenie_val,
            faza=detected.get('faza'),
            form_factor=detected.get('form_factor'),
            voltage=detected.get('voltage')
        )
        db.add(item)
        existing_serials.add(serial)
        count += 1
    
    register.items_count = count
    db.commit()
    return {
        "id": register.id, 
        "filename": register.filename, 
        "items_count": count, 
        "skipped_duplicates": skipped_duplicates,
        "duplicate_serials": duplicate_serials[:20],  # Первые 20 для показа
        "uploaded_at": register.uploaded_at
    }

@app.post("/api/pu/auto-fill-faza")
def auto_fill_faza(admin_code: str = Form(...), db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    """Массовое автозаполнение фазности по справочнику типов ПУ"""
    if not is_sue_admin(user):
        raise HTTPException(403, "Только СУЭ")
    if admin_code != settings.ADMIN_CODE:
        raise HTTPException(403, "Неверный код администратора")
    
    items = db.query(PUItem).filter(
        PUItem.faza == None,
        PUItem.pu_type != None
    ).all()
    
    updated = 0
    not_found_patterns = []
    for item in items:
        detected = detect_pu_type_params(item.pu_type, db)
        changed = False
        if detected.get('faza'):
            item.faza = detected['faza']
            changed = True
        if detected.get('form_factor') and not item.form_factor:
            item.form_factor = detected['form_factor']
            changed = True
        if detected.get('voltage') and not item.voltage:
            item.voltage = detected['voltage']
            changed = True
        if changed:
            updated += 1
        else:
            not_found_patterns.append(f"{item.serial_number} ({item.pu_type[:50]})")
    
    db.commit()
    return {
        "updated": updated, 
        "total_checked": len(items),
        "not_found_pu": not_found_patterns[:30]
    }

@app.post("/api/pu/import-naznachenie")
async def import_naznachenie(file: UploadFile = File(...), db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    """Массовая загрузка назначения ПУ из Excel (колонка 1 — номер ПУ, колонка 2 — назначение)"""
    if not is_sue_admin(user) and not is_lab_user(user):
        raise HTTPException(403, "Нет доступа")
    
    contents = await file.read()
    df = pd.read_excel(io.BytesIO(contents), header=None)
    
    updated = 0
    not_found = []
    errors = []
    naz_map = {'ижц': 'IZHC', 'техприс': 'TECHPRIS', 'замена': 'ZAMENA', 
               'izhc': 'IZHC', 'techpris': 'TECHPRIS', 'zamena': 'ZAMENA'}
    
    for idx, row in df.iterrows():
        serial = str(row.iloc[0]).strip()
        if not serial or serial == 'nan':
            continue
        
        naz_raw = str(row.iloc[1]).strip().lower() if len(row) > 1 and pd.notna(row.iloc[1]) else ''
        if not naz_raw or naz_raw == 'nan':
            continue
        
        naz_value = naz_map.get(naz_raw)
        if not naz_value:
            errors.append(f"{serial}: неизвестное назначение '{row.iloc[1]}'")
            continue
        
        item = db.query(PUItem).filter(PUItem.serial_number == serial).first()
        if not item:
            not_found.append(serial)
            continue
        
        item.naznachenie = naz_value
        updated += 1
    
    db.commit()
    return {
        "updated": updated,
        "not_found_count": len(not_found),
        "not_found": not_found[:20],
        "errors": errors[:20]
    }

@app.post("/api/pu/import-formfactor")
async def import_formfactor(file: UploadFile = File(...), admin_code: str = Form(...), db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    """Массовая загрузка форм-фактора ПУ (сплит/классика)"""
    if not is_sue_admin(user):
        raise HTTPException(403, "Только СУЭ")
    if admin_code != settings.ADMIN_CODE:
        raise HTTPException(403, "Неверный код администратора")
    
    contents = await file.read()
    df = pd.read_excel(io.BytesIO(contents), header=None)
    
    updated = 0
    not_found_pu = []
    errors = []
    total_rows = 0
    ff_map = {'сплит': 'split', 'классика': 'classic', 
              'split': 'split', 'classic': 'classic'}
    
    for idx, row in df.iterrows():
        serial = str(row.iloc[0]).strip()
        if not serial or serial == 'nan':
            continue
        total_rows += 1
        
        ff_raw = str(row.iloc[1]).strip().lower() if len(row) > 1 and pd.notna(row.iloc[1]) else ''
        if not ff_raw or ff_raw == 'nan':
            continue
        
        ff_value = ff_map.get(ff_raw)
        if not ff_value:
            errors.append(f"{serial}: неизвестный форм-фактор '{row.iloc[1]}'")
            continue
        
        item = db.query(PUItem).filter(PUItem.serial_number == serial).first()
        if not item:
            not_found_pu.append(serial)
            continue
        
        item.form_factor = ff_value
        updated += 1
    
    db.commit()
    return {
        "updated": updated,
        "total_rows": total_rows,
        "not_found_pu": not_found_pu[:20],
        "errors": errors[:20]
    }

@app.post("/api/pu/move")
def move_items(req: MoveReq, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    """Перемещение ПУ"""
    target = db.query(Unit).filter(Unit.id == req.to_unit_id).first()
    if not target:
        raise HTTPException(404, "Подразделение не найдено")
    
    items = db.query(PUItem).filter(PUItem.id.in_(req.pu_item_ids)).all()
    if not items:
        raise HTTPException(404, "ПУ не найдены")
    
    moved = 0
    for item in items:
        can_move, error = can_move_pu(user, item, target, db)
        if not can_move:
            raise HTTPException(403, error)
        
        mov = PUMovement(pu_item_id=item.id, from_unit_id=item.current_unit_id, to_unit_id=target.id, moved_by=user.id, comment=req.comment)
        db.add(mov)
        item.current_unit_id = target.id
        moved += 1
    
    db.commit()
    return {"moved": moved}

@app.post("/api/pu/move-bulk")
async def move_bulk(
    file: UploadFile = File(...),
    admin_code: str = Form(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user)
):
    """Массовое перемещение ПУ по Excel файлу (ЭСК / ОКС Админ)"""
    if not is_esk_admin(user) and not is_sue_admin(user) and not is_oks_admin(user):
        raise HTTPException(403, "Только ЭСК Админ, ОКС Админ или СУЭ")
    
    if admin_code != settings.ADMIN_CODE:
        raise HTTPException(403, "Неверный код администратора")
    
    try:
        contents = await file.read()
        df = pd.read_excel(io.BytesIO(contents), header=None)
        
        # Ищем заголовки или берём первые 2 колонки
        serial_col = 0
        unit_col = 1
        
        # Проверяем есть ли заголовок
        first_val = str(df.iloc[0, 0]).lower() if len(df) > 0 else ""
        start_row = 1 if 'номер' in first_val or 'серийн' in first_val or 'пу' in first_val else 0
        
        # Словарь подразделений: ОКС для ОКС-админа, иначе ЭСК
        if is_oks_admin(user):
            move_units = db.query(Unit).filter(Unit.unit_type.in_([UnitType.OKS, UnitType.OKS_UNIT])).all()
            strip_suffix = " РЭС"
        else:
            move_units = db.query(Unit).filter(Unit.unit_type.in_([UnitType.ESK, UnitType.ESK_UNIT])).all()
            strip_suffix = " ЭСК"
        units_map = {}
        for u in move_units:
            units_map[u.name.lower()] = u
            if u.code:
                units_map[u.code.lower()] = u
            # Короткие варианты: "Адлерский" -> подразделение
            short_name = u.name.replace(strip_suffix, "").lower()
            units_map[short_name] = u
        
        moved = 0
        not_found_pu = []
        not_found_unit = []
        errors = []
        
        for idx in range(start_row, len(df)):
            row = df.iloc[idx]
            serial = str(row.iloc[serial_col]).strip() if pd.notna(row.iloc[serial_col]) else ""
            unit_name = str(row.iloc[unit_col]).strip() if pd.notna(row.iloc[unit_col]) else ""
            
            if not serial or serial == 'nan' or not unit_name or unit_name == 'nan':
                continue
            
            # Ищем ПУ
            pu_item = db.query(PUItem).filter(PUItem.serial_number == serial).first()
            if not pu_item:
                not_found_pu.append(serial)
                continue
            
            # Ищем подразделение
            target_unit = units_map.get(unit_name.lower())
            if not target_unit:
                # Пробуем частичное совпадение
                for key, u in units_map.items():
                    if unit_name.lower() in key or key in unit_name.lower():
                        target_unit = u
                        break
            
            if not target_unit:
                not_found_unit.append(f"{serial}: {unit_name}")
                continue
            
            # Перемещаем
            try:
                mov = PUMovement(
                    pu_item_id=pu_item.id,
                    from_unit_id=pu_item.current_unit_id,
                    to_unit_id=target_unit.id,
                    moved_by=user.id,
                    comment=f"Массовое перемещение из файла {file.filename}"
                )
                db.add(mov)
                pu_item.current_unit_id = target_unit.id
                moved += 1
            except Exception as e:
                errors.append(f"{serial}: {str(e)}")
        
        db.commit()
        
        return {
            "moved": moved,
            "not_found_pu": not_found_pu,
            "not_found_unit": not_found_unit,
            "errors": errors,
            "total_rows": len(df) - start_row
        }
        
    except Exception as e:
        print(f"Move bulk error: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(500, f"Ошибка: {str(e)}")

@app.post("/api/pu/update-types-bulk")
async def update_types_bulk(
    file: UploadFile = File(...),
    admin_code: str = Form(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user)
):
    """Массовое обновление типов ПУ по Excel файлу"""
    if not is_sue_admin(user):
        raise HTTPException(403, "Только СУЭ")
    
    if admin_code != settings.ADMIN_CODE:
        raise HTTPException(403, "Неверный код администратора")
    
    try:
        contents = await file.read()
        
        # Читаем Excel с явными параметрами
        xl = pd.ExcelFile(io.BytesIO(contents))
        print(f"=== UPDATE TYPES BULK ===")
        print(f"Листы в файле: {xl.sheet_names}")
        
        # Читаем первый лист без заголовка, все строки
        df = pd.read_excel(xl, sheet_name=0, header=None, dtype=str)
        print(f"Всего строк в DataFrame: {len(df)}")
        print(f"Первые 5 строк: {df.head()}")
        
        # Проверяем есть ли заголовок
        first_val = str(df.iloc[0, 0]).lower() if len(df) > 0 else ""
        start_row = 1 if 'номер' in first_val or 'серийн' in first_val or 'пу' in first_val or 'заводской' in first_val else 0
        print(f"Начинаем с строки: {start_row}")
        
        updated = 0
        not_found = []
        errors = []
        
        for idx in range(start_row, len(df)):
            row = df.iloc[idx]
            
            # Получаем значения из первых двух колонок
            serial = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
            new_type = str(row.iloc[1]).strip() if len(row) > 1 and pd.notna(row.iloc[1]) else ""
            
            # Пропускаем пустые строки
            if not serial or serial == 'nan' or serial == 'None':
                continue
            
            # Убираем .0 если число было прочитано как float
            if serial.endswith('.0'):
                serial = serial[:-2]
            
            # Ищем ПУ
            pu_item = db.query(PUItem).filter(PUItem.serial_number == serial).first()
            if not pu_item:
                not_found.append(serial)
                continue
            
            if not new_type or new_type == 'nan' or new_type == 'None':
                errors.append(f"{serial}: пустой тип")
                continue
            
            # Обновляем тип
            pu_item.pu_type = new_type[:500]
            updated += 1
        
        db.commit()
        
        print(f"Обновлено: {updated}, Не найдено: {len(not_found)}, Ошибок: {len(errors)}")
        
        return {
            "updated": updated,
            "not_found": not_found,
            "errors": errors,
            "total_rows": len(df) - start_row
        }
        
    except Exception as e:
        print(f"Update types bulk error: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(500, f"Ошибка: {str(e)}")

@app.post("/api/pu/auto-fill-formfactor")
def auto_fill_formfactor(admin_code: str = Form(...), db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    """Массовое автозаполнение форм-фактора по справочнику типов ПУ"""
    if not is_sue_admin(user):
        raise HTTPException(403, "Только СУЭ")
    if admin_code != settings.ADMIN_CODE:
        raise HTTPException(403, "Неверный код администратора")
    
    items = db.query(PUItem).filter(
        PUItem.form_factor == None,
        PUItem.pu_type != None
    ).all()
    
    updated = 0
    not_found_patterns = []  # ПУ без совпадения по паттерну
    for item in items:
        detected = detect_pu_type_params(item.pu_type, db)
        if detected.get('form_factor'):
            item.form_factor = detected['form_factor']
            updated += 1
        else:
            not_found_patterns.append(f"{item.serial_number} ({item.pu_type[:50]})")
    
    db.commit()
    return {
        "updated": updated, 
        "total_checked": len(items),
        "not_found_pu": not_found_patterns[:30]
    }

@app.post("/api/pu/delete")
def delete_items(req: DeleteReq, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    """Удаление ПУ - только СУЭ с кодом"""
    if not is_sue_admin(user):
        raise HTTPException(403, "Только СУЭ может удалять ПУ")
    if req.admin_code != settings.ADMIN_CODE:
        raise HTTPException(403, "Неверный код администратора")
    
    # Удаляем связанные данные
    db.query(PUMovement).filter(PUMovement.pu_item_id.in_(req.pu_item_ids)).delete(synchronize_session=False)
    db.query(PUMaterial).filter(PUMaterial.pu_item_id.in_(req.pu_item_ids)).delete(synchronize_session=False)
    deleted = db.query(PUItem).filter(PUItem.id.in_(req.pu_item_ids)).delete(synchronize_session=False)
    
    db.commit()
    return {"deleted": deleted}

# ==================== ADMIN: БЭКАП И ДИАГНОСТИКА ====================

@app.post("/api/pu/clear-database")
def clear_database(data: dict, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    """Очистка базы данных - только СУЭ с кодом"""
    if not is_sue_admin(user):
        raise HTTPException(403, "Только СУЭ может очищать базу")
    if data.get("admin_code") != settings.ADMIN_CODE:
        raise HTTPException(403, "Неверный код администратора")
    
    db.query(PUMaterial).delete()
    db.query(PUMovement).delete()
    db.query(PUItem).delete()
    db.query(PURegister).delete()
    db.commit()
    
    return {"message": "База очищена"}

@app.get("/api/admin/backup")
def create_backup(admin_code: str, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    """Создать бэкап базы в JSON"""
    if admin_code != settings.ADMIN_CODE:
        raise HTTPException(403, "Неверный код")
    if not is_sue_admin(user):
        raise HTTPException(403, "Нет доступа")
    
    backup = {
        "created_at": datetime.now().isoformat(),
        "pu_items": [],
        "users": [],
        "units": [],
        "ttr_res": [],
        "ttr_esk": [],
        "materials": [],
        "va_nominals": [],
        "tt_nominals": [],
    }
    
    # ПУ
    for item in db.query(PUItem).all():
        backup["pu_items"].append({
            "id": item.id,
            "serial_number": item.serial_number,
            "pu_type": item.pu_type,
            "status": item.status.value if item.status else None,
            "current_unit_id": item.current_unit_id,
            "contract_number": item.contract_number,
            "consumer": item.consumer,
            "address": item.address,
            "faza": item.faza,
            "voltage": item.voltage,
            "power": item.power,
            "form_factor": item.form_factor,
            "trubostoyka": item.trubostoyka,
            "va_type": item.va_type,
            "has_va": item.has_va,
            "va_nominal_id": item.va_nominal_id,
            "has_tt": item.has_tt,
            "tt_nominal_id": item.tt_nominal_id,
            "approval_status": item.approval_status.value if item.approval_status else None,
            "tz_number": item.tz_number,
            "request_number": item.request_number,
        })
    
    # Номиналы ВА
    for item in db.query(VA_Nominal).filter(VA_Nominal.is_active == True).all():
        backup["va_nominals"].append({"id": item.id, "name": item.name})
    
    # Номиналы ТТ
    for item in db.query(TT_Nominal).filter(TT_Nominal.is_active == True).all():
        backup["tt_nominals"].append({"id": item.id, "name": item.name})
    
    # Материалы
    for item in db.query(Material).filter(Material.is_active == True).all():
        backup["materials"].append({"id": item.id, "name": item.name, "unit": item.unit})
    
    # ТТР РЭС
    for item in db.query(TTR_RES).filter(TTR_RES.is_active == True).all():
        backup["ttr_res"].append({
            "id": item.id, "code": item.code, "name": item.name, 
            "ttr_type": item.ttr_type, "use_tt": item.use_tt
        })
    
    # ТТР ЭСК
    for item in db.query(TTR_ESK).filter(TTR_ESK.is_active == True).all():
        backup["ttr_esk"].append({
            "id": item.id, "ttr_type": item.ttr_type, "work_type_name": item.work_type_name,
            "faza": item.faza, "form_factor": item.form_factor, "va_type": item.va_type,
            "lsr_number": item.lsr_number, "price_no_nds": item.price_no_nds, "price_with_nds": item.price_with_nds
        })
    
    # Возвращаем JSON файл
    output = io.BytesIO()
    output.write(json.dumps(backup, ensure_ascii=False, indent=2).encode('utf-8'))
    output.seek(0)
    
    filename = f"backup_{datetime.now().strftime('%Y%m%d_%H%M')}.json"
    
    return StreamingResponse(
        output,
        media_type="application/json",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@app.get("/api/admin/health-check")
def health_check(db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    """Проверка целостности базы"""
    if not is_sue_admin(user):
        raise HTTPException(403, "Нет доступа")
    
    issues = []
    
    # 1. ПУ без подразделения
    orphan_pu = db.query(PUItem).filter(PUItem.current_unit_id == None).count()
    if orphan_pu > 0:
        issues.append(f"⚠️ ПУ без подразделения: {orphan_pu}")
    
    # 2. ПУ с битыми ссылками на ТТР
    for item in db.query(PUItem).filter(PUItem.ttr_ou_id != None).all():
        ttr = db.query(TTR_RES).filter(TTR_RES.id == item.ttr_ou_id).first()
        if not ttr:
            issues.append(f"❌ ПУ {item.serial_number}: битая ссылка на ТТР ОУ (id={item.ttr_ou_id})")
    
    # 3. ПУ с ВА но без номинала
    va_without_nominal = db.query(PUItem).filter(
        PUItem.has_va == True, 
        PUItem.va_nominal_id == None
    ).count()
    if va_without_nominal > 0:
        issues.append(f"⚠️ ПУ с ВА но без номинала: {va_without_nominal}")
    
    # 4. ПУ с ТТ но без номинала
    tt_without_nominal = db.query(PUItem).filter(
        PUItem.has_tt == True, 
        PUItem.tt_nominal_id == None
    ).count()
    if tt_without_nominal > 0:
        issues.append(f"⚠️ ПУ с ТТ но без номинала: {tt_without_nominal}")
    
    # 5. Дубликаты серийных номеров
    from sqlalchemy import func
    duplicates = db.query(PUItem.serial_number, func.count(PUItem.id)).group_by(
        PUItem.serial_number
    ).having(func.count(PUItem.id) > 1).all()
    if duplicates:
        issues.append(f"❌ Дубликаты серийных номеров: {len(duplicates)}")
        for sn, cnt in duplicates[:5]:
            issues.append(f"   • {sn}: {cnt} шт")
    
    # 6. Дубликаты договоров
    dup_contracts = db.query(PUItem.contract_number, func.count(PUItem.id)).filter(
        PUItem.contract_number != None,
        PUItem.contract_number != ''
    ).group_by(PUItem.contract_number).having(func.count(PUItem.id) > 1).all()
    if dup_contracts:
        issues.append(f"⚠️ Дубликаты договоров: {len(dup_contracts)}")
    
    # Статистика
    stats = {
        "total_pu": db.query(PUItem).count(),
        "total_users": db.query(User).filter(User.is_active == True).count(),
        "total_ttr_res": db.query(TTR_RES).filter(TTR_RES.is_active == True).count(),
        "total_ttr_esk": db.query(TTR_ESK).filter(TTR_ESK.is_active == True).count(),
        "total_materials": db.query(Material).filter(Material.is_active == True).count(),
        "total_va_nominals": db.query(VA_Nominal).filter(VA_Nominal.is_active == True).count(),
        "total_tt_nominals": db.query(TT_Nominal).filter(TT_Nominal.is_active == True).count(),
    }
    
    return {
        "status": "OK" if len(issues) == 0 else "ISSUES_FOUND",
        "issues_count": len(issues),
        "issues": issues,
        "stats": stats,
        "checked_at": datetime.now().isoformat()
    }

@app.get("/api/admin/export-issues")
def export_issues_to_excel(db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    """Выгрузка ПУ с проблемами в Excel"""
    if not is_sue_admin(user):
        raise HTTPException(403, "Нет доступа")
    
    issues_data = []  # [{serial, pu_type, unit, status, problem, ...}]
    
    # 1. ПУ без подразделения
    for item in db.query(PUItem).filter(PUItem.current_unit_id == None).all():
        issues_data.append({"item": item, "problem": "Нет подразделения", "unit_name": "—"})
    
    # 2. Битые ссылки на ТТР
    for item in db.query(PUItem).options(joinedload(PUItem.current_unit)).filter(PUItem.ttr_ou_id != None).all():
        ttr = db.query(TTR_RES).filter(TTR_RES.id == item.ttr_ou_id).first()
        if not ttr:
            issues_data.append({"item": item, "problem": f"Битая ссылка ТТР ОУ (id={item.ttr_ou_id})", "unit_name": item.current_unit.name if item.current_unit else "—"})
    
    for item in db.query(PUItem).options(joinedload(PUItem.current_unit)).filter(PUItem.ttr_ol_id != None).all():
        ttr = db.query(TTR_RES).filter(TTR_RES.id == item.ttr_ol_id).first()
        if not ttr:
            issues_data.append({"item": item, "problem": f"Битая ссылка ТТР ОЛ (id={item.ttr_ol_id})", "unit_name": item.current_unit.name if item.current_unit else "—"})
    
    for item in db.query(PUItem).options(joinedload(PUItem.current_unit)).filter(PUItem.ttr_or_id != None).all():
        ttr = db.query(TTR_RES).filter(TTR_RES.id == item.ttr_or_id).first()
        if not ttr:
            issues_data.append({"item": item, "problem": f"Битая ссылка ТТР ОР (id={item.ttr_or_id})", "unit_name": item.current_unit.name if item.current_unit else "—"})
    
    # 3. ВА без номинала
    for item in db.query(PUItem).options(joinedload(PUItem.current_unit)).filter(PUItem.has_va == True, PUItem.va_nominal_id == None).all():
        issues_data.append({"item": item, "problem": "ВА без номинала", "unit_name": item.current_unit.name if item.current_unit else "—"})
    
    # 4. ТТ без номинала
    for item in db.query(PUItem).options(joinedload(PUItem.current_unit)).filter(PUItem.has_tt == True, PUItem.tt_nominal_id == None).all():
        issues_data.append({"item": item, "problem": "ТТ без номинала", "unit_name": item.current_unit.name if item.current_unit else "—"})
    
    # 5. Дубликаты серийных номеров
    from sqlalchemy import func as sql_func
    dup_serials = db.query(PUItem.serial_number).group_by(PUItem.serial_number).having(sql_func.count(PUItem.id) > 1).all()
    dup_sn_set = {d[0] for d in dup_serials}
    if dup_sn_set:
        for item in db.query(PUItem).options(joinedload(PUItem.current_unit)).filter(PUItem.serial_number.in_(dup_sn_set)).all():
            issues_data.append({"item": item, "problem": "Дубликат серийного номера", "unit_name": item.current_unit.name if item.current_unit else "—"})
    
    # 6. Дубликаты договоров
    dup_contracts = db.query(PUItem.contract_number).filter(
        PUItem.contract_number != None, PUItem.contract_number != ''
    ).group_by(PUItem.contract_number).having(sql_func.count(PUItem.id) > 1).all()
    dup_cn_set = {d[0] for d in dup_contracts}
    if dup_cn_set:
        for item in db.query(PUItem).options(joinedload(PUItem.current_unit)).filter(PUItem.contract_number.in_(dup_cn_set)).all():
            # Не дублируем если уже есть по другой причине
            if not any(e["item"].id == item.id and "Дубликат договора" in e["problem"] for e in issues_data):
                issues_data.append({"item": item, "problem": f"Дубликат договора ({item.contract_number})", "unit_name": item.current_unit.name if item.current_unit else "—"})
    
    # 7. Техприс без договора
    for item in db.query(PUItem).options(joinedload(PUItem.current_unit)).filter(
        PUItem.status == PUStatus.TECHPRIS,
        or_(PUItem.contract_number == None, PUItem.contract_number == '')
    ).all():
        issues_data.append({"item": item, "problem": "Техприс без договора", "unit_name": item.current_unit.name if item.current_unit else "—"})
    
    # 8. Замена/ИЖЦ без ЛС
    for item in db.query(PUItem).options(joinedload(PUItem.current_unit)).filter(
        PUItem.status.in_([PUStatus.ZAMENA, PUStatus.IZHC]),
        or_(PUItem.ls_number == None, PUItem.ls_number == '')
    ).all():
        issues_data.append({"item": item, "problem": f"{'Замена' if item.status == PUStatus.ZAMENA else 'ИЖЦ'} без ЛС", "unit_name": item.current_unit.name if item.current_unit else "—"})
    
    # 9. Не на складе без фазности
    for item in db.query(PUItem).options(joinedload(PUItem.current_unit)).filter(
        PUItem.status != PUStatus.SKLAD,
        or_(PUItem.faza == None, PUItem.faza == '')
    ).all():
        issues_data.append({"item": item, "problem": "Нет фазности", "unit_name": item.current_unit.name if item.current_unit else "—"})
    
    # 10. Не на складе без ТТР ОУ
    for item in db.query(PUItem).options(joinedload(PUItem.current_unit)).filter(
        PUItem.status != PUStatus.SKLAD,
        PUItem.ttr_ou_id == None,
        PUItem.current_unit.has(Unit.unit_type == UnitType.RES)
    ).all():
        issues_data.append({"item": item, "problem": "Нет ТТР орг. учета", "unit_name": item.current_unit.name if item.current_unit else "—"})
    
    if not issues_data:
        raise HTTPException(404, "Проблем не найдено! Все ПУ в порядке.")
    
    # Создаём Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Проблемные ПУ"
    
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    headers = [
        ("№", 5),
        ("Проблема", 35),
        ("Серийный номер", 20),
        ("Тип ПУ", 35),
        ("Подразделение", 20),
        ("Статус", 12),
        ("Договор", 22),
        ("ЛС", 15),
        ("ТЗ", 15),
        ("Заявка", 12),
        ("Фазность", 10),
        ("ТТР ОУ", 10),
        ("ТТР ОЛ", 10),
        ("ТТР ОР", 10),
    ]
    
    for col, (header, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col)].width = width
    
    ws.row_dimensions[1].height = 35
    
    status_labels = {'SKLAD': 'Склад', 'TECHPRIS': 'Техприс', 'ZAMENA': 'Замена', 'IZHC': 'ИЖЦ'}
    
    # Убираем дубли (один ПУ может иметь несколько проблем — группируем)
    pu_problems = {}
    for entry in issues_data:
        item = entry["item"]
        if item.id not in pu_problems:
            pu_problems[item.id] = {"item": item, "problems": [], "unit_name": entry["unit_name"]}
        pu_problems[item.id]["problems"].append(entry["problem"])
    
    warn_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    err_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
    
    for idx, (pu_id, data) in enumerate(pu_problems.items(), 1):
        row = idx + 1
        item = data["item"]
        problems_str = "; ".join(data["problems"])
        is_error = any(w in problems_str for w in ["Дубликат", "Битая"])
        fill = err_fill if is_error else warn_fill
        
        row_data = [
            idx,
            problems_str,
            item.serial_number or "",
            item.pu_type or "",
            data["unit_name"],
            status_labels.get(item.status.value, item.status.value) if item.status else "",
            item.contract_number or "",
            item.ls_number or "",
            item.tz_number or "",
            item.request_number or "",
            item.faza or "",
            item.ttr_ou_id or "",
            item.ttr_ol_id or "",
            item.ttr_or_id or "",
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = thin_border
            cell.fill = fill
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        
        ws.row_dimensions[row].height = 28
    
    # Итого
    total_row = len(pu_problems) + 2
    ws.cell(row=total_row, column=1, value=f"Всего проблемных ПУ: {len(pu_problems)}")
    ws.cell(row=total_row, column=1).font = Font(bold=True)
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename_utf8 = f"Проблемные_ПУ_{datetime.now().strftime('%d.%m.%Y')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename_utf8)}"}
    )
@app.post("/api/admin/restore")
def restore_backup(
    file: UploadFile = File(...),
    admin_code: str = None,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user)
):
    """Восстановить базу из JSON бэкапа"""
    if not admin_code or admin_code != settings.ADMIN_CODE:
        raise HTTPException(403, "Неверный код администратора")
    if not is_sue_admin(user):
        raise HTTPException(403, "Нет доступа")
    
    try:
        content = file.file.read()
        backup = json.loads(content.decode('utf-8'))
    except Exception as e:
        raise HTTPException(400, f"Ошибка чтения файла: {str(e)}")
    
    restored = {
        "va_nominals": 0,
        "tt_nominals": 0,
        "materials": 0,
        "ttr_res": 0,
        "ttr_esk": 0,
        "pu_items": 0,
    }
    
    # 1. Восстанавливаем номиналы ВА
    for item in backup.get("va_nominals", []):
        existing = db.query(VA_Nominal).filter(VA_Nominal.id == item["id"]).first()
        if not existing:
            db.add(VA_Nominal(id=item["id"], name=item["name"], is_active=True))
            restored["va_nominals"] += 1
    
    # 2. Восстанавливаем номиналы ТТ
    for item in backup.get("tt_nominals", []):
        existing = db.query(TT_Nominal).filter(TT_Nominal.id == item["id"]).first()
        if not existing:
            db.add(TT_Nominal(id=item["id"], name=item["name"], is_active=True))
            restored["tt_nominals"] += 1
    
    # 3. Восстанавливаем материалы
    for item in backup.get("materials", []):
        existing = db.query(Material).filter(Material.id == item["id"]).first()
        if not existing:
            db.add(Material(id=item["id"], name=item["name"], unit=item["unit"], is_active=True))
            restored["materials"] += 1
    
    # 4. Восстанавливаем ТТР РЭС
    for item in backup.get("ttr_res", []):
        existing = db.query(TTR_RES).filter(TTR_RES.id == item["id"]).first()
        if not existing:
            db.add(TTR_RES(
                id=item["id"], code=item["code"], name=item["name"],
                ttr_type=item["ttr_type"], use_tt=item.get("use_tt", False), is_active=True
            ))
            restored["ttr_res"] += 1
    
    # 5. Восстанавливаем ТТР ЭСК
    for item in backup.get("ttr_esk", []):
        existing = db.query(TTR_ESK).filter(TTR_ESK.id == item["id"]).first()
        if not existing:
            db.add(TTR_ESK(
                id=item["id"], ttr_type=item.get("ttr_type"), work_type_name=item.get("work_type_name"),
                faza=item.get("faza"), form_factor=item.get("form_factor"), va_type=item.get("va_type"),
                lsr_number=item.get("lsr_number"), price_no_nds=item.get("price_no_nds"),
                price_with_nds=item.get("price_with_nds"), is_active=True
            ))
            restored["ttr_esk"] += 1
    
    db.commit()
    
    # 6. Восстанавливаем ПУ
    for item in backup.get("pu_items", []):
        existing = db.query(PUItem).filter(PUItem.serial_number == item["serial_number"]).first()
        if not existing:
            pu = PUItem(
                serial_number=item["serial_number"],
                pu_type=item.get("pu_type"),
                status=PUStatus(item["status"]) if item.get("status") else PUStatus.SKLAD,
                current_unit_id=item.get("current_unit_id"),
                contract_number=item.get("contract_number"),
                consumer=item.get("consumer"),
                address=item.get("address"),
                faza=item.get("faza"),
                voltage=item.get("voltage"),
                power=item.get("power"),
                form_factor=item.get("form_factor"),
                trubostoyka=item.get("trubostoyka"),
                va_type=item.get("va_type"),
                has_va=item.get("has_va", False),
                va_nominal_id=item.get("va_nominal_id"),
                has_tt=item.get("has_tt", False),
                tt_nominal_id=item.get("tt_nominal_id"),
                approval_status=ApprovalStatus(item["approval_status"]) if item.get("approval_status") else None,
                tz_number=item.get("tz_number"),
                request_number=item.get("request_number"),
            )
            db.add(pu)
            restored["pu_items"] += 1
    
    db.commit()
    
    return {
        "status": "OK",
        "message": "Восстановление завершено",
        "restored": restored
    }

# ==================== API: СОГЛАСОВАНИЕ (ЭСК -> РЭС) ====================
@app.post("/api/pu/items/{item_id}/send-approval")
def send_for_approval(item_id: int, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    """Отправить на согласование (ЭСК / ОКС)"""
    if not (is_esk_user(user) or is_esk_admin(user) or is_oks_user(user) or is_oks_admin(user)):
        raise HTTPException(403, "Только ЭСК или ОКС может отправлять на согласование")
    
    item = db.query(PUItem).filter(PUItem.id == item_id).first()
    if not item:
        raise HTTPException(404, "ПУ не найден")
    
    item.approval_status = ApprovalStatus.PENDING
    item.rejection_comment = None  # Сбрасываем комментарий при повторной отправке
    db.commit()
    return {"ok": True}

@app.post("/api/pu/send-approval-batch")
def send_approval_batch(data: dict, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    """Массовая отправка на согласование (ЭСК / ОКС)"""
    if not (is_esk_user(user) or is_esk_admin(user) or is_oks_user(user) or is_oks_admin(user)):
        raise HTTPException(403, "Только ЭСК или ОКС может отправлять на согласование")
    
    item_ids = data.get("item_ids", [])
    if not item_ids:
        raise HTTPException(400, "Не выбраны ПУ")
    
    updated = db.query(PUItem).filter(
        PUItem.id.in_(item_ids),
        PUItem.approval_status != ApprovalStatus.APPROVED  # Не трогаем уже согласованные
    ).update({"approval_status": ApprovalStatus.PENDING}, synchronize_session=False)
    
    db.commit()
    return {"updated": updated}

@app.post("/api/pu/items/{item_id}/approve")
def approve_item(item_id: int, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    """Согласовать (РЭС)"""
    if not is_res_user(user) and not is_sue_admin(user):
        raise HTTPException(403, "Только РЭС может согласовывать")
    
    item = db.query(PUItem).filter(PUItem.id == item_id).first()
    if not item:
        raise HTTPException(404, "ПУ не найден")
    
    item.approval_status = ApprovalStatus.APPROVED
    item.approved_by = user.id
    item.approved_at = datetime.utcnow()
    db.commit()
    return {"ok": True}

@app.post("/api/pu/items/{item_id}/reject")
def reject_item(item_id: int, data: dict, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    """Отклонить ПУ (РЭС) с комментарием"""
    if not is_res_user(user) and not is_sue_admin(user):
        raise HTTPException(403, "Только РЭС может отклонять")
    
    item = db.query(PUItem).filter(PUItem.id == item_id).first()
    if not item:
        raise HTTPException(404, "ПУ не найден")
    
    comment = data.get("comment", "").strip()
    if not comment:
        raise HTTPException(400, "Укажите причину отклонения")
    
    item.approval_status = ApprovalStatus.REJECTED
    item.rejection_comment = comment
    item.approved_by = user.id
    item.approved_at = datetime.utcnow()
    db.commit()
    return {"ok": True}

@app.post("/api/pu/items/{item_id}/unlock")
def unlock_item(item_id: int, data: dict, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    """Разблокировать согласованную карточку (только СУЭ с кодом)"""
    if not is_sue_admin(user):
        raise HTTPException(403, "Только СУЭ может разблокировать")
    if data.get("admin_code") != settings.ADMIN_CODE:
        raise HTTPException(403, "Неверный код администратора")
    
    item = db.query(PUItem).filter(PUItem.id == item_id).first()
    if not item:
        raise HTTPException(404, "ПУ не найден")
    
    item.approval_status = ApprovalStatus.NONE
    item.approved_by = None
    item.approved_at = None
    db.commit()
    return {"ok": True}


@app.get("/api/pu/pending-approval")
def get_pending_approval(db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    """Список ПУ на согласовании для РЭС"""
    if not is_res_user(user) and not is_sue_admin(user):
        raise HTTPException(403, "Нет доступа")
    
    if is_res_user(user) and user.unit:
        base = user.unit.code.replace("RES_", "") if user.unit.code else ""
        target_codes = [f"ESK_{base}", f"OKS_{base}"]
        target_ids = [u.id for u in db.query(Unit).filter(Unit.code.in_(target_codes)).all()]
        if target_ids:
            items = db.query(PUItem).options(joinedload(PUItem.current_unit)).filter(
                PUItem.current_unit_id.in_(target_ids),
                PUItem.approval_status == ApprovalStatus.PENDING
            ).all()
        else:
            items = []
    else:
        items = db.query(PUItem).options(joinedload(PUItem.current_unit)).filter(PUItem.approval_status == ApprovalStatus.PENDING).all()
    
    code_to_name = {u.code: u.name for u in db.query(Unit.code, Unit.name).all()}
    
    def get_res_name(src_unit):
        if not src_unit or not src_unit.code:
            return "—"
        res_code = src_unit.code.replace("ESK_", "RES_").replace("OKS_", "RES_")
        return code_to_name.get(res_code, "—")
    
    def get_source(src_unit):
        if src_unit and src_unit.unit_type in (UnitType.OKS, UnitType.OKS_UNIT):
            return "ОКС"
        return "ЭСК"
    
    return [{
        "id": i.id, 
        "serial_number": i.serial_number, 
        "pu_type": i.pu_type,
        "current_unit_name": i.current_unit.name if i.current_unit else None,
        "res_name": get_res_name(i.current_unit),
        "source": get_source(i.current_unit),
        "contract_number": i.contract_number, 
        "consumer": i.consumer,
        "address": i.address,
        "faza": i.faza,
        "form_factor": i.form_factor,
        "trubostoyka": i.trubostoyka,
        "va_type": i.va_type,
        "lsr_va": i.lsr_va,
        "lsr_truba": i.lsr_truba,
        "smr_date": i.smr_date.isoformat() if i.smr_date else None,
    } for i in items]

@app.get("/api/pu/pending-approval/export")
def export_pending_approval(db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    """Выгрузка реестра на согласовании в Excel"""
    if not is_res_user(user) and not is_sue_admin(user):
        raise HTTPException(403, "Нет доступа")
    
    if is_res_user(user) and user.unit:
        base = user.unit.code.replace("RES_", "") if user.unit.code else ""
        target_codes = [f"ESK_{base}", f"OKS_{base}"]
        target_ids = [u.id for u in db.query(Unit).filter(Unit.code.in_(target_codes)).all()]
        if target_ids:
            items = db.query(PUItem).filter(
                PUItem.current_unit_id.in_(target_ids),
                PUItem.approval_status == ApprovalStatus.PENDING
            ).all()
        else:
            items = []
    else:
        items = db.query(PUItem).filter(PUItem.approval_status == ApprovalStatus.PENDING).all()
    
    code_to_name = {u.code: u.name for u in db.query(Unit.code, Unit.name).all()}
    
    def get_res_name(src_unit):
        if not src_unit or not src_unit.code:
            return "—"
        res_code = src_unit.code.replace("ESK_", "RES_").replace("OKS_", "RES_")
        return code_to_name.get(res_code, "—")
    
    def get_source(src_unit):
        if src_unit and src_unit.unit_type in (UnitType.OKS, UnitType.OKS_UNIT):
            return "ОКС"
        return "ЭСК"
    
    # Создаём Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "На согласовании"
    
    # Стили
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Заголовки
    headers = [
        ("№", 5),
        ("Источник", 10),
        ("РЭС", 18),
        ("Серийный номер", 20),
        ("Тип ПУ", 40),
        ("Потребитель", 25),
        ("Адрес", 35),
        ("Договор", 22),
        ("Фазность", 10),
        ("Трубостойка", 12),
        ("№ ТТР ЭСК", 15),
        ("Дата СМР", 12),
    ]
    
    for col, (header, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col)].width = width
    
    ws.row_dimensions[1].height = 35
    
    # Данные
    for idx, item in enumerate(items, 1):
        row = idx + 1
        data = [
            idx,
            get_source(item.current_unit),
            get_res_name(item.current_unit),
            item.serial_number or "",
            item.pu_type or "",
            item.consumer or "",
            item.address or "",
            item.contract_number or "",
            item.faza or "",
            "Да" if item.trubostoyka else "Нет",
            item.lsr_va or item.lsr_truba or "",
            item.smr_date.strftime("%d.%m.%Y") if item.smr_date else "",
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        
        ws.row_dimensions[row].height = 25
    
    # Сохраняем
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename_rus = f"На_согласовании_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=\"pending_approval.xlsx\"; filename*=UTF-8''{quote(filename_rus)}"
        }
    )

# ==================== API: СПРАВОЧНИКИ (CRUD) ====================

# --- Мастера ЭСК ---
@app.post("/api/masters")
def create_master(data: dict, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    if not is_esk_admin(user) and not is_sue_admin(user):
        raise HTTPException(403, "Нет доступа")
    master = ESKMaster(full_name=data["full_name"], unit_id=data["unit_id"])
    db.add(master)
    db.commit()
    return {"id": master.id}

@app.put("/api/masters/{master_id}")
def update_master(master_id: int, data: dict, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    if not is_esk_admin(user) and not is_sue_admin(user):
        raise HTTPException(403, "Нет доступа")
    m = db.query(ESKMaster).filter(ESKMaster.id == master_id).first()
    if not m:
        raise HTTPException(404, "Не найден")
    for k, v in data.items():
        if hasattr(m, k):
            setattr(m, k, v)
    db.commit()
    return {"ok": True}

@app.delete("/api/masters/{master_id}")
def delete_master(master_id: int, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    if not is_esk_admin(user) and not is_sue_admin(user):
        raise HTTPException(403, "Нет доступа")
    db.query(ESKMaster).filter(ESKMaster.id == master_id).delete()
    db.commit()
    return {"ok": True}

# --- ТТР РЭС ---
@app.post("/api/ttr/res")
def create_ttr_res(data: dict, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    if not is_sue_admin(user):
        raise HTTPException(403, "Нет доступа")
    ttr = TTR_RES(code=data["code"], name=data["name"], ttr_type=data["ttr_type"], pu_types=data.get("pu_types", ""))
    db.add(ttr)
    db.commit()
    return {"id": ttr.id}

@app.put("/api/ttr/res/{ttr_id}")
def update_ttr_res(ttr_id: int, data: dict, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    if not is_sue_admin(user):
        raise HTTPException(403, "Нет доступа")
    t = db.query(TTR_RES).filter(TTR_RES.id == ttr_id).first()
    if not t:
        raise HTTPException(404, "Не найден")
    for k, v in data.items():
        if hasattr(t, k):
            setattr(t, k, v)
    db.commit()
    return {"ok": True}

@app.delete("/api/ttr/res/{ttr_id}")
def delete_ttr_res(ttr_id: int, data: dict = None, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    if not is_sue_admin(user):
        raise HTTPException(403, "Нет доступа")
    if not data or data.get("admin_code") != settings.ADMIN_CODE:
        raise HTTPException(403, "Неверный код администратора")
    
    # Удаляем связанные материалы
    db.query(TTR_Material).filter(TTR_Material.ttr_res_id == ttr_id).delete()
    db.query(TTR_RES).filter(TTR_RES.id == ttr_id).delete()
    db.commit()
    return {"ok": True}

@app.delete("/api/materials/{mat_id}")
def delete_material(mat_id: int, data: dict = None, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    if not is_sue_admin(user):
        raise HTTPException(403, "Нет доступа")
    if not data or data.get("admin_code") != settings.ADMIN_CODE:
        raise HTTPException(403, "Неверный код администратора")
    
    # Удаляем связи с ТТР и ПУ
    db.query(TTR_Material).filter(TTR_Material.material_id == mat_id).delete()
    db.query(PUMaterial).filter(PUMaterial.material_id == mat_id).delete()
    db.query(Material).filter(Material.id == mat_id).delete()
    db.commit()
    return {"ok": True}

# ==================== API: СПРАВОЧНИКИ ВА и ТТ ====================

@app.get("/api/va-nominals")
def get_va_nominals(db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    items = db.query(VA_Nominal).filter(VA_Nominal.is_active == True).all()
    return [{"id": v.id, "name": v.name} for v in items]


@app.post("/api/va-nominals")
def create_va_nominal(data: dict, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    if not is_sue_admin(user):
        raise HTTPException(403, "Нет доступа")
    v = VA_Nominal(name=data["name"])
    db.add(v)
    db.commit()
    return {"id": v.id}


@app.put("/api/va-nominals/{item_id}")
def update_va_nominal(item_id: int, data: dict, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    if not is_sue_admin(user):
        raise HTTPException(403, "Нет доступа")
    v = db.query(VA_Nominal).filter(VA_Nominal.id == item_id).first()
    if not v:
        raise HTTPException(404, "Не найден")
    for k, val in data.items():
        if hasattr(v, k):
            setattr(v, k, val)
    db.commit()
    return {"ok": True}


@app.delete("/api/va-nominals/{item_id}")
def delete_va_nominal(item_id: int, data: dict = None, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    if not is_sue_admin(user):
        raise HTTPException(403, "Нет доступа")
    if not data or data.get("admin_code") != settings.ADMIN_CODE:
        raise HTTPException(403, "Неверный код администратора")
    db.query(VA_Nominal).filter(VA_Nominal.id == item_id).update({"is_active": False})
    db.commit()
    return {"ok": True}


@app.get("/api/tt-nominals")
def get_tt_nominals(db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    items = db.query(TT_Nominal).filter(TT_Nominal.is_active == True).all()
    return [{"id": t.id, "name": t.name} for t in items]


@app.post("/api/tt-nominals")
def create_tt_nominal(data: dict, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    if not is_sue_admin(user):
        raise HTTPException(403, "Нет доступа")
    t = TT_Nominal(name=data["name"])
    db.add(t)
    db.commit()
    return {"id": t.id}


@app.put("/api/tt-nominals/{item_id}")
def update_tt_nominal(item_id: int, data: dict, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    if not is_sue_admin(user):
        raise HTTPException(403, "Нет доступа")
    t = db.query(TT_Nominal).filter(TT_Nominal.id == item_id).first()
    if not t:
        raise HTTPException(404, "Не найден")
    for k, val in data.items():
        if hasattr(t, k):
            setattr(t, k, val)
    db.commit()
    return {"ok": True}


@app.delete("/api/tt-nominals/{item_id}")
def delete_tt_nominal(item_id: int, data: dict = None, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    if not is_sue_admin(user):
        raise HTTPException(403, "Нет доступа")
    if not data or data.get("admin_code") != settings.ADMIN_CODE:
        raise HTTPException(403, "Неверный код администратора")
    db.query(TT_Nominal).filter(TT_Nominal.id == item_id).update({"is_active": False})
    db.commit()
    return {"ok": True}

# --- ТТР ЭСК ---
@app.post("/api/ttr/esk")
def create_ttr_esk(data: dict, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    if not is_sue_admin(user):
        raise HTTPException(403, "Нет доступа")
    ttr = TTR_ESK(
        faza=data.get("faza"),
        form_factor=data.get("form_factor"),
        va_type=data.get("va_type"),
        lsr_number=data.get("lsr_number"),
        price_no_nds=data.get("price_no_nds", 0),
        price_with_nds=data.get("price_with_nds", 0)
    )
    db.add(ttr)
    db.commit()
    return {"id": ttr.id}

@app.put("/api/ttr/esk/{ttr_id}")
def update_ttr_esk(ttr_id: int, data: dict, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    if not is_sue_admin(user):
        raise HTTPException(403, "Нет доступа")
    t = db.query(TTR_ESK).filter(TTR_ESK.id == ttr_id).first()
    if not t:
        raise HTTPException(404, "Не найден")
    for k, v in data.items():
        if hasattr(t, k):
            setattr(t, k, v)
    db.commit()
    return {"ok": True}

@app.delete("/api/ttr/esk/{ttr_id}")
def delete_ttr_esk(ttr_id: int, data: dict = None, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    if not is_sue_admin(user):
        raise HTTPException(403, "Нет доступа")
    if not data or data.get("admin_code") != settings.ADMIN_CODE:
        raise HTTPException(403, "Неверный код администратора")
    db.query(TTR_ESK).filter(TTR_ESK.id == ttr_id).update({"is_active": False})
    db.commit()
    return {"ok": True}

# --- Материалы ---
@app.get("/api/materials")
def get_materials(db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    items = db.query(Material).filter(Material.is_active == True).all()
    return [{"id": m.id, "name": m.name, "unit": m.unit} for m in items]

@app.post("/api/materials")
def create_material(data: dict, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    if not is_sue_admin(user):
        raise HTTPException(403, "Нет доступа")
    m = Material(name=data["name"], unit=data.get("unit", "шт"))
    db.add(m)
    db.commit()
    return {"id": m.id}

@app.put("/api/materials/{mat_id}")
def update_material(mat_id: int, data: dict, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    if not is_sue_admin(user):
        raise HTTPException(403, "Нет доступа")
    m = db.query(Material).filter(Material.id == mat_id).first()
    if not m:
        raise HTTPException(404, "Не найден")
    for k, v in data.items():
        if hasattr(m, k):
            setattr(m, k, v)
    db.commit()
    return {"ok": True}

# --- Материалы к ТТР ---
@app.get("/api/ttr/res/{ttr_id}/materials")
def get_ttr_materials(ttr_id: int, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    items = db.query(TTR_Material).filter(TTR_Material.ttr_res_id == ttr_id).all()
    result = []
    for tm in items:
        mat = db.query(Material).filter(Material.id == tm.material_id).first()
        if mat:
            result.append({"id": tm.id, "material_id": mat.id, "material_name": mat.name, "unit": mat.unit, "quantity": tm.quantity})
    return result

@app.post("/api/ttr/res/{ttr_id}/materials")
def set_ttr_materials(ttr_id: int, data: dict, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    if not is_sue_admin(user):
        raise HTTPException(403, "Нет доступа")
    # data = {"materials": [{"material_id": 1, "quantity": 5}, ...]}
    db.query(TTR_Material).filter(TTR_Material.ttr_res_id == ttr_id).delete()
    for m in data.get("materials", []):
        tm = TTR_Material(ttr_res_id=ttr_id, material_id=m["material_id"], quantity=m["quantity"])
        db.add(tm)
    db.commit()
    return {"ok": True}

@app.get("/api/ttr/res/{ttr_id}/pu-types")
def get_ttr_pu_types(ttr_id: int, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    """Получить привязанные типы ПУ к ТТР"""
    items = db.query(TTR_PUType).filter(TTR_PUType.ttr_res_id == ttr_id).all()
    result = []
    for item in items:
        pu_type = db.query(PUTypeReference).filter(PUTypeReference.id == item.pu_type_id).first()
        if pu_type:
            result.append({
                "id": item.id,
                "pu_type_id": pu_type.id,
                "pattern": pu_type.pattern,
                "faza": pu_type.faza,
                "voltage": pu_type.voltage
            })
    return result


@app.post("/api/ttr/res/{ttr_id}/pu-types")
def set_ttr_pu_types(ttr_id: int, data: dict, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    """Установить привязку типов ПУ к ТТР"""
    if not is_sue_admin(user):
        raise HTTPException(403, "Нет доступа")
    
    # Удаляем старые связи
    db.query(TTR_PUType).filter(TTR_PUType.ttr_res_id == ttr_id).delete()
    
    # Добавляем новые
    for pu_type_id in data.get("pu_type_ids", []):
        link = TTR_PUType(ttr_res_id=ttr_id, pu_type_id=pu_type_id)
        db.add(link)
    
    db.commit()
    return {"ok": True}


@app.get("/api/ttr/res/for-pu")
def get_ttr_for_pu(pu_type: str, ttr_type: str, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    """Получить ТТР доступные для данного типа ПУ"""
    print(f"=== TTR FOR PU ===")
    print(f"pu_type: {pu_type}")
    print(f"ttr_type: {ttr_type}")
    
    if not pu_type:
        print("pu_type пустой — возвращаем []")
        return []
    
    pu_type_upper = pu_type.upper().strip()
    
    # Ищем подходящий тип ПУ из справочника
    all_pu_types = db.query(PUTypeReference).filter(PUTypeReference.is_active == True).all()
    print(f"Всего типов ПУ в справочнике: {len(all_pu_types)}")
    
    matched_pu_type = None
    for pt in sorted(all_pu_types, key=lambda x: len(x.pattern or ''), reverse=True):
        if pt.pattern and pt.pattern.upper() in pu_type_upper:
            print(f"  ✓ СОВПАЛ: '{pt.pattern}' в '{pu_type_upper}'")
            matched_pu_type = pt
            break
    
    if not matched_pu_type:
        print("Тип ПУ не найден — возвращаем []")
        return []
    
    print(f"Найден тип ПУ: id={matched_pu_type.id}, pattern={matched_pu_type.pattern}")
    
    # Ищем ТТР привязанные к этому типу ПУ
    linked = db.query(TTR_PUType).filter(TTR_PUType.pu_type_id == matched_pu_type.id).all()
    print(f"Записей в TTR_PUType для pu_type_id={matched_pu_type.id}: {len(linked)}")
    
    linked_ttr_ids = [l.ttr_res_id for l in linked]
    
    if not linked_ttr_ids:
        print("Нет привязанных ТТР — возвращаем []")
        return []
    
    # Фильтруем по типу ТТР (OU, OL, OR)
    ttrs = db.query(TTR_RES).filter(
        TTR_RES.id.in_(linked_ttr_ids),
        TTR_RES.ttr_type == ttr_type,
        TTR_RES.is_active == True
    ).all()
    
    print(f"Найдено ТТР типа '{ttr_type}': {len(ttrs)}")
    for t in ttrs:
        print(f"  - id={t.id}, code={t.code}, ttr_type={t.ttr_type}")
    
    # ⭐ ИСПРАВЛЕНИЕ: добавляем ttr_type и use_tt в ответ!
    return [{"id": t.id, "code": t.code, "name": t.name, "ttr_type": t.ttr_type, "use_tt": t.use_tt} for t in ttrs]

@app.get("/api/pu/items/{item_id}/materials")
def get_pu_materials(
    item_id: int, 
    ttr_ou_id: Optional[int] = None,
    ttr_ol_id: Optional[int] = None, 
    ttr_or_id: Optional[int] = None,
    ttr_tt_id: Optional[int] = None,
    db: Session = Depends(get_db), 
    user: User = Depends(get_current_user)
):
    """Получить материалы для ПУ (из выбранных ТТР)"""
    item = db.query(PUItem).filter(PUItem.id == item_id).first()
    if not item:
        raise HTTPException(404, "ПУ не найден")
    
    # Если переданы параметры - используем их, иначе берём из сохранённого item
    if ttr_ou_id is not None or ttr_ol_id is not None or ttr_or_id is not None or ttr_tt_id is not None:
        ttr_ids = [t for t in [ttr_ou_id, ttr_ol_id, ttr_or_id, ttr_tt_id] if t]
    else:
        ttr_ids = [t for t in [item.ttr_ou_id, item.ttr_ol_id, item.ttr_or_id, item.ttr_tt_id] if t]
    
    if not ttr_ids:
        return {"defaults": [], "facts": []}
    
    # Материалы по умолчанию из ТТР (суммируем)
    defaults = {}
    for ttr_id in ttr_ids:
        ttr_mats = db.query(TTR_Material).filter(TTR_Material.ttr_res_id == ttr_id).all()
        for tm in ttr_mats:
            mat = db.query(Material).filter(Material.id == tm.material_id).first()
            if mat:
                key = mat.id
                if key in defaults:
                    defaults[key]["quantity"] += tm.quantity
                else:
                    defaults[key] = {
                        "material_id": mat.id,
                        "material_name": mat.name,
                        "unit": mat.unit,
                        "quantity": tm.quantity
                    }
    
    # Фактические значения (если уже заполняли)
    facts = db.query(PUMaterial).filter(PUMaterial.pu_item_id == item_id).all()
    facts_list = []
    for f in facts:
        mat = db.query(Material).filter(Material.id == f.material_id).first()
        if mat:
            facts_list.append({
                "id": f.id,
                "material_id": mat.id,
                "material_name": mat.name,
                "unit": mat.unit,
                "quantity": f.quantity,
                "used": f.used
            })
    
    return {
        "defaults": list(defaults.values()),
        "facts": facts_list
    }


@app.post("/api/pu/items/{item_id}/materials")
def save_pu_materials(item_id: int, data: dict, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    """Сохранить фактические материалы для ПУ (РЭС и ОКС работают с материалами)"""
    if not is_res_user(user) and not is_sue_admin(user) and not is_oks_user(user):
        raise HTTPException(403, "Нет доступа")
    
    item = db.query(PUItem).filter(PUItem.id == item_id).first()
    if not item:
        raise HTTPException(404, "ПУ не найден")
    
    # Удаляем старые записи
    db.query(PUMaterial).filter(PUMaterial.pu_item_id == item_id).delete()
    
    # Добавляем новые
    # data = {"materials": [{"material_id": 1, "quantity": 5, "used": true}, ...]}
    for m in data.get("materials", []):
        pm = PUMaterial(
            pu_item_id=item_id,
            material_id=m["material_id"],
            quantity=m.get("quantity", 0),
            used=m.get("used", True)
        )
        db.add(pm)
    
    item.materials_used = True
    db.commit()
    return {"ok": True}

@app.post("/api/pu/items/materials-bulk")
def get_materials_bulk(data: dict, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    """Получить материалы для нескольких ПУ"""
    item_ids = data.get("item_ids", [])
    if not item_ids:
        return []
    
    result = []
    for item_id in item_ids:
        item = db.query(PUItem).filter(PUItem.id == item_id).first()
        if not item:
            continue
        
        ttr_ids = [t for t in [item.ttr_ou_id, item.ttr_ol_id, item.ttr_or_id, item.ttr_tt_id] if t]
        
        # Материалы по умолчанию из ТТР
        defaults = {}
        for ttr_id in ttr_ids:
            ttr_mats = db.query(TTR_Material).filter(TTR_Material.ttr_res_id == ttr_id).all()
            for tm in ttr_mats:
                mat = db.query(Material).filter(Material.id == tm.material_id).first()
                if mat:
                    key = mat.id
                    if key in defaults:
                        defaults[key]["quantity"] += tm.quantity
                    else:
                        defaults[key] = {
                            "material_id": mat.id,
                            "material_name": mat.name,
                            "unit": mat.unit,
                            "quantity": tm.quantity
                        }
        
        # Фактические значения
        facts = db.query(PUMaterial).filter(PUMaterial.pu_item_id == item_id).all()
        if facts:
            materials = []
            for f in facts:
                mat = db.query(Material).filter(Material.id == f.material_id).first()
                if mat:
                    materials.append({
                        "material_id": mat.id,
                        "material_name": mat.name,
                        "unit": mat.unit,
                        "quantity": f.quantity,
                        "used": f.used
                    })
        else:
            materials = [{"material_id": d["material_id"], "material_name": d["material_name"], 
                          "unit": d["unit"], "quantity": d["quantity"], "used": True} 
                         for d in defaults.values()]
        
        result.append({
            "id": item.id,
            "serial_number": item.serial_number,
            "pu_type": item.pu_type,
            "ttr_ou": item.ttr_ou.code if item.ttr_ou else None,
            "ttr_ol": item.ttr_ol.code if item.ttr_ol else None,
            "ttr_or": item.ttr_or.code if item.ttr_or else None,
            "ttr_tt": item.ttr_tt.code if item.ttr_tt else None,
            "has_va": item.has_va,
            "va_nominal_name": item.va_nominal.name if item.va_nominal else None,
            "va_quantity": item.va_quantity or 1,
            "has_tt": item.has_tt,
            "tt_nominal_name": item.tt_nominal.name if item.tt_nominal else None,
            "materials": materials
        })
    
    return result


@app.post("/api/pu/items/materials-bulk/save")
def save_materials_bulk(data: dict, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    """Сохранить материалы для нескольких ПУ"""
    if not is_res_user(user) and not is_sue_admin(user):
        raise HTTPException(403, "Нет доступа")
    
    items_data = data.get("items", [])
    saved = 0
    
    for item_data in items_data:
        item_id = item_data.get("item_id")
        materials = item_data.get("materials", [])
    
        item = db.query(PUItem).filter(PUItem.id == item_id).first()
        if not item:
            continue
    
        # Удаляем старые записи
        db.query(PUMaterial).filter(PUMaterial.pu_item_id == item_id).delete()
    
    # Добавляем новые
        for m in materials:
                pm = PUMaterial(
                    pu_item_id=item_id,
                    material_id=m["material_id"],
                    quantity=m.get("quantity", 0),
                    used=m.get("used", True)
                )
                db.add(pm)
    
    # Сохраняем ВА и ТТ
        va_used = item_data.get("va_used")
        va_quantity = item_data.get("va_quantity")
        tt_used = item_data.get("tt_used")
    
        if va_used is not None:
            item.has_va = va_used
            if not va_used:
                item.ttr_or_id = None
                item.va_nominal_id = None
        if va_quantity is not None:
            item.va_quantity = va_quantity
        if tt_used is not None:
            item.has_tt = tt_used
    
        item.materials_used = True
        saved += 1
    
    db.commit()
    return {"saved": saved}

# --- Справочник типов ПУ ---
@app.get("/api/pu-types")
def get_pu_types(db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    items = db.query(PUTypeReference).filter(PUTypeReference.is_active == True).all()
    return [{"id": p.id, "pattern": p.pattern, "faza": p.faza, "voltage": p.voltage, "form_factor": p.form_factor} for p in items]

@app.post("/api/pu-types")
def create_pu_type(data: dict, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    if not is_sue_admin(user):
        raise HTTPException(403, "Нет доступа")
    p = PUTypeReference(
        pattern=data["pattern"], 
        faza=data.get("faza"), 
        voltage=data.get("voltage"),
        form_factor=data.get("form_factor")
    )
    db.add(p)
    db.commit()
    return {"id": p.id}

@app.put("/api/pu-types/{type_id}")
def update_pu_type(type_id: int, data: dict, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    if not is_sue_admin(user):
        raise HTTPException(403, "Нет доступа")
    p = db.query(PUTypeReference).filter(PUTypeReference.id == type_id).first()
    if not p:
        raise HTTPException(404, "Не найден")
    for k, v in data.items():
        if hasattr(p, k):
            setattr(p, k, v)
    db.commit()
    return {"ok": True}

@app.delete("/api/pu-types/{type_id}")
def delete_pu_type(type_id: int, data: dict = None, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    if not is_sue_admin(user):
        raise HTTPException(403, "Нет доступа")
    if not data or data.get("admin_code") != settings.ADMIN_CODE:
        raise HTTPException(403, "Неверный код администратора")
    db.query(PUTypeReference).filter(PUTypeReference.id == type_id).update({"is_active": False})
    db.commit()
    return {"ok": True}

# ==================== API: ТЗ и ЗАЯВКИ ====================

@app.get("/api/tz/list")
def get_tz_list(tz_type: Optional[str] = None, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    """Список ТЗ"""
    q = db.query(PUItem).options(joinedload(PUItem.current_unit)).filter(PUItem.tz_number != None, PUItem.tz_number != "")
    if tz_type:
        q = q.filter(PUItem.status == tz_type)
    
    # Группируем по номеру ТЗ
    items = q.all()
    tz_map = {}
    for item in items:
        if item.tz_number not in tz_map:
            tz_map[item.tz_number] = {
                "tz_number": item.tz_number,
                "status": item.status.value,
                "unit_name": item.current_unit.name if item.current_unit else None,
                "count": 0,
                "items": []
            }
        tz_map[item.tz_number]["count"] += 1
        tz_map[item.tz_number]["items"].append(item.id)
    
    return list(tz_map.values())

@app.get("/api/tz/export")
def export_tz_to_excel(tz_number: str = Query(...), db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    """Выгрузка ТЗ в Excel с материалами"""
    try:
        items = db.query(PUItem).filter(PUItem.tz_number == tz_number).all()
        
        if not items:
            raise HTTPException(404, "ТЗ не найден")
        
        # Создаём книгу Excel
        wb = openpyxl.Workbook()
        
        # ===== ЛИСТ 1: Список ПУ =====
        ws1 = wb.active
        ws1.title = "Список ПУ"
        
        # Стили
        header_font = Font(bold=True, color="FFFFFF", size=10)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # Заголовки листа 1
        headers1 = [
            ("№", 5),
            ("Серийный номер", 20),
            ("Тип ПУ", 40),
            ("ЛС", 15),
            ("Потребитель", 25),
            ("Адрес", 35),
            ("Договор", 22),
            ("Мощность", 10),
            ("Фазность", 10),
            ("Напряжение", 12),
            ("ТТР ОУ", 12),
            ("ТТР ОЛ", 12),
            ("ТТР ОР", 12),
            ("ТТР ТТ", 12),
            ("ВА", 10),
            ("ТТ", 10),
        ]
        
        for col, (header, width) in enumerate(headers1, 1):
            cell = ws1.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            ws1.column_dimensions[get_column_letter(col)].width = width
        
        ws1.row_dimensions[1].height = 35
        
        # Данные листа 1
        for idx, item in enumerate(items, 1):
            row = idx + 1
            data = [
                idx,
                item.serial_number or "",
                item.pu_type or "",
                item.ls_number or "",
                item.consumer or "",
                item.address or "",
                item.contract_number or "",
                item.power or "",
                item.faza or "",
                item.voltage or "",
                item.ttr_ou.code if item.ttr_ou else "",
                item.ttr_ol.code if item.ttr_ol else "",
                item.ttr_or.code if item.ttr_or else "",
                item.ttr_tt.code if item.ttr_tt else "",
                item.va_nominal.name if item.va_nominal else "",
                item.tt_nominal.name if item.tt_nominal else "",
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws1.cell(row=row, column=col, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center", wrap_text=True)
            
            ws1.row_dimensions[row].height = 25
        
        # ===== ЛИСТ 2: Материалы по каждому ПУ =====
        ws2 = wb.create_sheet("Материалы по ПУ")
        
        headers2 = [
            ("№", 5),
            ("Серийный номер", 20),
            ("Тип ПУ", 30),
            ("ТТР", 25),
            ("Материал", 30),
            ("Ед.", 8),
            ("Кол-во", 10),
        ]
        
        for col, (header, width) in enumerate(headers2, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            ws2.column_dimensions[get_column_letter(col)].width = width
        
        row_num = 2
        for idx, item in enumerate(items, 1):
            # Получаем материалы для этого ПУ
            pu_materials = db.query(PUMaterial).filter(
                PUMaterial.pu_item_id == item.id,
                PUMaterial.used == True
            ).all()
            
            # Если нет сохранённых материалов — берём из ТТР
            if not pu_materials:
                ttr_ids = [t for t in [item.ttr_ou_id, item.ttr_ol_id, item.ttr_or_id, item.ttr_tt_id] if t]
                materials_dict = {}
                for ttr_id in ttr_ids:
                    ttr_mats = db.query(TTR_Material).filter(TTR_Material.ttr_res_id == ttr_id).all()
                    for tm in ttr_mats:
                        mat = db.query(Material).filter(Material.id == tm.material_id).first()
                        if mat:
                            if mat.id in materials_dict:
                                materials_dict[mat.id]['quantity'] += tm.quantity
                            else:
                                materials_dict[mat.id] = {
                                    'name': mat.name,
                                    'unit': mat.unit,
                                    'quantity': tm.quantity
                                }
                
                for mat_data in materials_dict.values():
                    ttr_codes = ", ".join([t.code for t in [item.ttr_ou, item.ttr_ol, item.ttr_or, item.ttr_tt] if t])
                    data = [
                        idx,
                        item.serial_number or "",
                        item.pu_type or "",
                        ttr_codes,
                        mat_data['name'],
                        mat_data['unit'],
                        mat_data['quantity'],
                    ]
                    for col, value in enumerate(data, 1):
                        cell = ws2.cell(row=row_num, column=col, value=value)
                        cell.border = thin_border
                    row_num += 1
            else:
                for pm in pu_materials:
                    mat = db.query(Material).filter(Material.id == pm.material_id).first()
                    if mat:
                        ttr_codes = ", ".join([t.code for t in [item.ttr_ou, item.ttr_ol, item.ttr_or, item.ttr_tt] if t])
                        data = [
                            idx,
                            item.serial_number or "",
                            item.pu_type or "",
                            ttr_codes,
                            mat.name,
                            mat.unit,
                            pm.quantity,
                        ]
                        for col, value in enumerate(data, 1):
                            cell = ws2.cell(row=row_num, column=col, value=value)
                            cell.border = thin_border
                        row_num += 1
        
        # ===== ЛИСТ 3: Сводная по материалам =====
        ws3 = wb.create_sheet("Сводная материалов")
        
        headers3 = [
            ("№", 5),
            ("Материал", 40),
            ("Ед. изм.", 10),
            ("Всего", 12),
        ]
        
        for col, (header, width) in enumerate(headers3, 1):
            cell = ws3.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            ws3.column_dimensions[get_column_letter(col)].width = width
        
        # Собираем сводную
        totals = {}
        for item in items:
            pu_materials = db.query(PUMaterial).filter(
                PUMaterial.pu_item_id == item.id,
                PUMaterial.used == True
            ).all()
            
            if pu_materials:
                for pm in pu_materials:
                    mat = db.query(Material).filter(Material.id == pm.material_id).first()
                    if mat:
                        if mat.id not in totals:
                            totals[mat.id] = {'name': mat.name, 'unit': mat.unit, 'quantity': 0}
                        totals[mat.id]['quantity'] += pm.quantity
            else:
                # Из ТТР
                ttr_ids = [t for t in [item.ttr_ou_id, item.ttr_ol_id, item.ttr_or_id, item.ttr_tt_id] if t]
                for ttr_id in ttr_ids:
                    ttr_mats = db.query(TTR_Material).filter(TTR_Material.ttr_res_id == ttr_id).all()
                    for tm in ttr_mats:
                        mat = db.query(Material).filter(Material.id == tm.material_id).first()
                        if mat:
                            if mat.id not in totals:
                                totals[mat.id] = {'name': mat.name, 'unit': mat.unit, 'quantity': 0}
                            totals[mat.id]['quantity'] += tm.quantity
        
        # ВА и ТТ в сводную
        va_totals = {}
        tt_totals = {}
        for item in items:
            if item.has_va and item.va_nominal:
                name = f"ВА {item.va_nominal.name}"
                va_totals[name] = va_totals.get(name, 0) + 1
            if item.has_tt and item.tt_nominal:
                name = f"ТТ {item.tt_nominal.name}"
                tt_totals[name] = tt_totals.get(name, 0) + 1
        
        row_num = 2
        for idx, (mat_id, mat_data) in enumerate(totals.items(), 1):
            data = [idx, mat_data['name'], mat_data['unit'], mat_data['quantity']]
            for col, value in enumerate(data, 1):
                cell = ws3.cell(row=row_num, column=col, value=value)
                cell.border = thin_border
            row_num += 1
        
        # Добавляем ВА
        for name, count in va_totals.items():
            data = [row_num - 1, name, 'шт', count]
            for col, value in enumerate(data, 1):
                cell = ws3.cell(row=row_num, column=col, value=value)
                cell.border = thin_border
                if col == 2:
                    cell.font = Font(bold=True, color="B45F06")
            row_num += 1
        
        # Добавляем ТТ
        for name, count in tt_totals.items():
            data = [row_num - 1, name, 'шт', count]
            for col, value in enumerate(data, 1):
                cell = ws3.cell(row=row_num, column=col, value=value)
                cell.border = thin_border
                if col == 2:
                    cell.font = Font(bold=True, color="7030A0")
            row_num += 1
        
        # Итоговая строка
        ws3.cell(row=row_num + 1, column=1, value=f"Всего ПУ: {len(items)} шт.")
        ws3.cell(row=row_num + 1, column=1).font = Font(bold=True)
        
        # Сохраняем
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
       # Безопасное имя файла (ASCII для filename=, UTF-8 для filename*=)
        safe_tz_ascii = re.sub(r'[^a-zA-Z0-9_\-]', '_', tz_number)  # Только ASCII
        filename_rus = f"ТЗ_{tz_number.replace('/', '-')}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename=\"TZ_{safe_tz_ascii}.xlsx\"; filename*=UTF-8''{quote(filename_rus)}"
            }
        )
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"Export TZ error: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(500, f"Ошибка экспорта: {str(e)}")

@app.get("/api/tz/{tz_number}/items")
def get_tz_items(tz_number: str, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    """Получить все ПУ по номеру ТЗ"""
    items = db.query(PUItem).options(joinedload(PUItem.current_unit)).filter(PUItem.tz_number == tz_number).all()
    return [{
     "id": i.id,
     "serial_number": i.serial_number,
     "pu_type": i.pu_type,
     "status": i.status.value,
     "current_unit_name": i.current_unit.name if i.current_unit else None,
     "contract_number": i.contract_number,
     "consumer": i.consumer,
     "address": i.address,
     "power": i.power,
     "faza": i.faza,
     "voltage": i.voltage,
     "ls_number": i.ls_number,
     "ttr_ou_id": i.ttr_ou_id,
     "ttr_ol_id": i.ttr_ol_id,
     "ttr_or_id": i.ttr_or_id,
    } for i in items]



@app.get("/api/tz/pending")
def get_pending_for_tz(
    status: str, 
    unit_id: Optional[int] = None, 
    power_category: Optional[int] = None,
    db: Session = Depends(get_db), 
    user: User = Depends(get_current_user)
):
    """ПУ без ТЗ для формирования с фильтром по мощности"""
    if not is_sue_admin(user) and not is_oks_admin(user):
        raise HTTPException(403, "Только СУЭ или ОКС может формировать ТЗ")
    
    q = db.query(PUItem).options(joinedload(PUItem.current_unit)).filter(
    PUItem.status == status,
    (PUItem.tz_number == None) | (PUItem.tz_number == "")
    )
    
    if unit_id:
        q = q.filter(PUItem.current_unit_id == unit_id)
    
    # Фильтр по категории мощности ТОЛЬКО для Техприс

    if status == 'TECHPRIS' and power_category:
        if power_category == 1:
            q = q.filter((PUItem.power == None) | (PUItem.power <= 15))
        elif power_category == 2:
            q = q.filter(PUItem.power > 15, PUItem.power <= 150)
        elif power_category == 3:
            q = q.filter(PUItem.power > 150)
    
    # СУЭ берёт только ПУ РЭС, ОКС — только ПУ ОКС
    if is_oks_admin(user):
        scope_units = db.query(Unit.id).filter(Unit.unit_type.in_([UnitType.OKS, UnitType.OKS_UNIT]))
    else:
        scope_units = db.query(Unit.id).filter(Unit.unit_type == UnitType.RES)
    q = q.filter(PUItem.current_unit_id.in_(scope_units))
    
    items = q.all()
    return [{
        "id": i.id, "serial_number": i.serial_number, "pu_type": i.pu_type,
        "current_unit_name": i.current_unit.name if i.current_unit else None,
        "current_unit_id": i.current_unit_id,
        "power": i.power
    } for i in items]

@app.post("/api/tz/create")
def create_tz(data: dict, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    """Создать ТЗ с автоматическим номером"""
    if not is_sue_admin(user) and not is_oks_admin(user):
        raise HTTPException(403, "Только СУЭ или ОКС может формировать ТЗ")
    
    item_ids = data["item_ids"]
    unit_id = data["unit_id"]  # РЭС (для СУЭ) или участок ОКС (для ОКС)
    status = data["status"]  # TECHPRIS, ZAMENA, IZHC
    power_category = data.get("power_category")  # Только для Техприс
    custom_suffix = data.get("custom_suffix")  # Ручная корректировка окончания
    
    # Получаем букву РЭС
    unit = db.query(Unit).filter(Unit.id == unit_id).first()
    if not unit or not unit.short_code:
        raise HTTPException(400, "Подразделение не найдено или не указан код")
    
    # Проверка соответствия типа подразделения роли
    is_oks_unit = unit.unit_type in (UnitType.OKS, UnitType.OKS_UNIT)
    if is_oks_admin(user) and not is_oks_unit:
        raise HTTPException(403, "ОКС может формировать ТЗ только по своим подразделениям")
    if is_sue_admin(user) and unit.unit_type != UnitType.RES:
        raise HTTPException(403, "СУЭ может формировать ТЗ только по РЭС")
    
    # Формируем префикс в зависимости от типа
    if status == 'TECHPRIS':
        prefix = f"ТП {power_category}"
    elif status == 'ZAMENA':
        prefix = "522"
    elif status == 'IZHC':
        prefix = "ИЖЦ"
    else:
        prefix = status
    
    # Формируем номер
    now = datetime.utcnow()
    if custom_suffix:
        suffix = custom_suffix
    else:
        month = now.strftime("%m")
        year = now.strftime("%y")
        suffix = f"{month}-{year}"
    
    # Для ОКС добавляем маркер в конце номера (после даты), чтобы отличать
    # от ТЗ РЭС и избежать коллизий номеров
    oks_suffix = "-ОКС" if is_oks_unit else ""
    tz_number = f"{prefix} {unit.short_code}-{suffix}{oks_suffix}"
    
    # Проверяем уникальность
    existing = db.query(PUItem).filter(PUItem.tz_number == tz_number).first()
    if existing:
        raise HTTPException(400, f"ТЗ с номером {tz_number} уже существует")
    
    updated = db.query(PUItem).filter(PUItem.id.in_(item_ids)).update({"tz_number": tz_number}, synchronize_session=False)
    db.commit()
    
    return {"created": updated, "tz_number": tz_number}

@app.post("/api/tz/remove-items")
def remove_items_from_tz(data: dict, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    """Удалить (исключить) ПУ из ТЗ — СУЭ или ОКС"""
    if not is_sue_admin(user) and not is_oks_admin(user):
        raise HTTPException(403, "Только СУЭ или ОКС может удалять ПУ из ТЗ")
    
    item_ids = data.get("item_ids", [])
    tz_number = data.get("tz_number", "")
    
    if not item_ids:
        raise HTTPException(400, "Не выбраны ПУ для удаления из ТЗ")
    
    # Обновляем только ПУ, которые принадлежат указанному ТЗ
    rm_q = db.query(PUItem).filter(
        PUItem.id.in_(item_ids),
        PUItem.tz_number == tz_number
    )
    # ОКС может исключать только ПУ своих подразделений
    if is_oks_admin(user):
        oks_units = db.query(Unit.id).filter(Unit.unit_type.in_([UnitType.OKS, UnitType.OKS_UNIT]))
        rm_q = rm_q.filter(PUItem.current_unit_id.in_(oks_units))
    updated = rm_q.update({"tz_number": None}, synchronize_session=False)
    
    db.commit()
    
    # Проверяем, остались ли ещё ПУ в этом ТЗ
    remaining = db.query(PUItem).filter(PUItem.tz_number == tz_number).count()
    
    return {"removed": updated, "remaining": remaining, "tz_number": tz_number}

@app.get("/api/tz/search-available")
def search_available_for_tz(
    tz_number: str,
    q: str = "",
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user)
):
    """Поиск ПУ для добавления в существующий ТЗ"""
    if not is_sue_admin(user) and not is_oks_admin(user):
        raise HTTPException(403, "Только СУЭ или ОКС может добавлять ПУ в ТЗ")
    
    if not q or len(q) < 2:
        return []
    
    # Определяем тип ТЗ по его номеру, чтобы фильтровать по статусу
    sample = db.query(PUItem).filter(PUItem.tz_number == tz_number).first()
    if not sample:
        raise HTTPException(404, "ТЗ не найден или пуст")
    
    tz_status = sample.status.value  # TECHPRIS, ZAMENA, IZHC
    
    # Ищем ПУ без ТЗ с подходящим статусом
    query = db.query(PUItem).filter(
        PUItem.tz_number.is_(None),
        PUItem.status == tz_status,
    )
    
    # Фильтр по подразделению (только из того же подразделения что и ТЗ)
    if sample.current_unit_id:
        query = query.filter(PUItem.current_unit_id == sample.current_unit_id)
    
    # ОКС может добавлять только ПУ своих подразделений
    if is_oks_admin(user):
        oks_units = db.query(Unit.id).filter(Unit.unit_type.in_([UnitType.OKS, UnitType.OKS_UNIT]))
        query = query.filter(PUItem.current_unit_id.in_(oks_units))
    
    # Поиск по серийному номеру или договору/ЛС
    search_filter = or_(
        PUItem.serial_number.ilike(f"%{q}%"),
        PUItem.contract_number.ilike(f"%{q}%"),
        PUItem.ls_number.ilike(f"%{q}%"),
    )
    query = query.filter(search_filter)
    
    items = query.limit(20).all()
    
    return [{
        "id": i.id,
        "serial_number": i.serial_number,
        "pu_type": i.pu_type,
        "status": i.status.value,
        "consumer": i.consumer,
        "address": i.address,
        "ls_number": i.ls_number,
        "contract_number": i.contract_number,
        "power": i.power,
        "faza": i.faza,
    } for i in items]

@app.post("/api/tz/add-items")
def add_items_to_tz(data: dict, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    """Добавить ПУ в существующий ТЗ — СУЭ или ОКС"""
    if not is_sue_admin(user) and not is_oks_admin(user):
        raise HTTPException(403, "Только СУЭ или ОКС может добавлять ПУ в ТЗ")
    
    item_ids = data.get("item_ids", [])
    tz_number = data.get("tz_number", "")
    
    if not item_ids:
        raise HTTPException(400, "Не выбраны ПУ для добавления")
    if not tz_number:
        raise HTTPException(400, "Не указан номер ТЗ")
    
    # Проверяем что ТЗ существует
    existing_count = db.query(PUItem).filter(PUItem.tz_number == tz_number).count()
    if existing_count == 0:
        raise HTTPException(404, "ТЗ не найден")
    
    # Добавляем только ПУ без ТЗ
    add_q = db.query(PUItem).filter(
        PUItem.id.in_(item_ids),
        PUItem.tz_number.is_(None)
    )
    # ОКС может добавлять только ПУ своих подразделений
    if is_oks_admin(user):
        oks_units = db.query(Unit.id).filter(Unit.unit_type.in_([UnitType.OKS, UnitType.OKS_UNIT]))
        add_q = add_q.filter(PUItem.current_unit_id.in_(oks_units))
    updated = add_q.update({"tz_number": tz_number}, synchronize_session=False)
    
    db.commit()
    
    total = db.query(PUItem).filter(PUItem.tz_number == tz_number).count()
    
    return {"added": updated, "total": total, "tz_number": tz_number}

@app.get("/api/tz/next-number")
def get_next_tz_number(
    status: str,
    unit_id: int,
    power_category: Optional[int] = None,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user)
):
    """Получить следующий номер ТЗ"""
    unit = db.query(Unit).filter(Unit.id == unit_id).first()
    if not unit or not unit.short_code:
        return {"next_suffix": "", "preview": "—"}
    
    now = datetime.utcnow()
    month = now.strftime("%m")
    year = now.strftime("%y")
    
    # Формируем префикс
    if status == 'TECHPRIS':
        prefix = f"ТП {power_category}"
    elif status == 'ZAMENA':
        prefix = "522"
    elif status == 'IZHC':
        prefix = "ИЖЦ"
    else:
        prefix = status
    
    # Ищем последний номер с таким префиксом
    pattern = f"{prefix} {unit.short_code}/%"
    last_tz = db.query(PUItem).filter(
        PUItem.tz_number.like(pattern)
    ).order_by(PUItem.id.desc()).first()
    
    next_suffix = f"{month}-{year}"
    
    # Маркер ОКС в конце номера для участков ОКС
    oks_suffix = "-ОКС" if unit.unit_type in (UnitType.OKS, UnitType.OKS_UNIT) else ""
    
    return {
        "next_suffix": next_suffix,
        "preview": f"{prefix} {unit.short_code}-{next_suffix}{oks_suffix}"
    }

@app.get("/api/requests/list")
def get_requests_list(db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    """Список заявок ЭСК"""
    q = db.query(PUItem).options(joinedload(PUItem.current_unit)).filter(PUItem.request_number != None, PUItem.request_number != "")
    
    # ЭСК видит только свои заявки
    if is_esk_user(user) or is_esk_admin(user):
        visible = get_visible_units(user, db)
        q = q.filter(PUItem.current_unit_id.in_(visible))
    
    items = q.all()
    req_map = {}
    for item in items:
        key = f"{item.request_number}|{item.request_contract or ''}"
        if key not in req_map:
            req_map[key] = {
                "request_number": item.request_number,
                "request_contract": item.request_contract,
                "display_name": f"№ {item.request_number} Договор № {item.request_contract}" if item.request_contract else f"№ {item.request_number}",
                "unit_name": item.current_unit.name if item.current_unit else None,
                "count": 0
            }
        req_map[key]["count"] += 1
    
    return list(req_map.values())

@app.get("/api/requests/{request_number}/items")
def get_request_items(request_number: str, request_contract: Optional[str] = None, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    """Получить все ПУ по номеру заявки с расширенными данными"""
    q = db.query(PUItem).options(joinedload(PUItem.current_unit)).filter(PUItem.request_number == request_number)
    if request_contract:
        q = q.filter(PUItem.request_contract == request_contract)
    
    items = q.all()
    
    # Получаем связанные РЭС для каждого ЭСК
    def get_res_name(esk_unit):
        if not esk_unit:
            return "—"
        res_code = esk_unit.code.replace("ESK_", "RES_") if esk_unit.code else ""
        res_unit = db.query(Unit).filter(Unit.code == res_code).first()
        return res_unit.name if res_unit else "—"
    
    return [{
        "id": i.id,
        "row_num": idx + 1,
        "filial": "Сочинский ПЭС",
        "res_name": get_res_name(i.current_unit),
        "serial_number": i.serial_number,
        "pu_type": i.pu_type,
        "consumer": i.consumer,
        "address": i.address,
        "contract_number": i.contract_number,
        "contract_date": i.contract_date.isoformat() if i.contract_date else None,
        "plan_date": i.plan_date.isoformat() if i.plan_date else None,
        "power": i.power,
        "faza": i.faza,
        "work_type_name": i.work_type_name,
        "price_with_nds": (i.price_truba_with_nds or 0) + (i.price_va_with_nds or 0),
        "current_unit_name": i.current_unit.name if i.current_unit else None,
    } for idx, i in enumerate(items)]

def get_res_name_for_esk(esk_unit, db):
    """Получить название РЭС для подразделения ЭСК"""
    if not esk_unit or not esk_unit.code:
        return "—"
    res_code = esk_unit.code.replace("ESK_", "RES_")
    res_unit = db.query(Unit).filter(Unit.code == res_code).first()
    return res_unit.name if res_unit else "—"

@app.get("/api/requests/pending")
def get_pending_for_request(unit_id: Optional[int] = None, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    """Согласованные ПУ для заявки ЭСК"""
    if not is_esk_admin(user) and not is_esk_user(user):
        raise HTTPException(403, "Только ЭСК может формировать заявки")
    
    q = db.query(PUItem).options(joinedload(PUItem.current_unit)).filter(
    PUItem.approval_status == ApprovalStatus.APPROVED,
    (PUItem.request_number == None) | (PUItem.request_number == "")
    )
    
    visible = get_visible_units(user, db)
    q = q.filter(PUItem.current_unit_id.in_(visible))
    
    if unit_id:
        q = q.filter(PUItem.current_unit_id == unit_id)
    
    items = q.all()
    return [{
        "id": i.id, 
        "res_name": get_res_name_for_esk(i.current_unit, db),
        "serial_number": i.serial_number, 
        "pu_type": i.pu_type,
        "current_unit_name": i.current_unit.name if i.current_unit else None,
        "contract_number": i.contract_number, 
        "consumer": i.consumer,
        "address": i.address,
        "faza": i.faza,
        "form_factor": i.form_factor,
        "trubostoyka": i.trubostoyka,
        "va_type": i.va_type,
        "lsr_truba": i.lsr_truba,
        "lsr_va": i.lsr_va,
        "price_truba_with_nds": i.price_truba_with_nds,
        "price_va_with_nds": i.price_va_with_nds,
        "price_truba_no_nds": i.price_truba_no_nds,
        "price_va_no_nds": i.price_va_no_nds,
        "price_total": (i.price_truba_with_nds or 0) + (i.price_va_with_nds or 0),
        "work_type_name": i.work_type_name,
    } for i in items]

@app.get("/api/requests/{request_number}/export")
def export_request_to_excel(
    request_number: str, 
    request_contract: Optional[str] = None, 
    db: Session = Depends(get_db), 
    user: User = Depends(get_current_user)
):
    """Выгрузка заявки в Excel"""
    try:
        q = db.query(PUItem).options(joinedload(PUItem.current_unit)).filter(PUItem.request_number == request_number)
        if request_contract:
            q = q.filter(PUItem.request_contract == request_contract)
        
        items = q.all()
        
        if not items:
            raise HTTPException(404, "Заявка не найдена")
        
        # Создаём книгу Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        safe_title = re.sub(r'[\\/?*:\[\]]', '-', f"Заявка {request_number}")
        ws.title = safe_title[:31]  # Excel ограничивает имя листа 31 символом
        
        # Стили
        header_font = Font(bold=True, color="FFFFFF", size=10)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        cell_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        center_alignment = Alignment(horizontal="center", vertical="center")
        money_alignment = Alignment(horizontal="right", vertical="center")
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Заголовки
        headers = [
            ("№", 5),
            ("Филиал", 20),
            ("РЭС", 18),
            ("Заявитель (ФИО)", 25),
            ("Адрес объекта", 35),
            ("Номер договора", 22),
            ("Дата заключения", 14),
            ("План. дата", 14),
            ("Мощность", 10),
            ("Тип ПУ", 35),
            ("Номер ПУ", 18),
            ("Фазность", 10),
            ("Вид работ", 25),
            ("ЛСР ПУ/ВА", 12),
            ("С НДС", 12),
            ("Трубост.", 10),
            ("ЛСР Труб.", 12),
            ("С НДС", 12),
            ("ИТОГО с НДС", 14),
        ]
        
        # Записываем заголовки
        for col, (header, width) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # Высота заголовка
        ws.row_dimensions[1].height = 40
        
        # Предзагружаем все РЭС единым запросом (вместо 130 отдельных)
        res_units = {u.code: u.name for u in db.query(Unit).filter(Unit.unit_type == UnitType.RES).all()}
        
        def get_res_name(esk_unit):
            if not esk_unit or not esk_unit.code:
                return "—"
            res_code = esk_unit.code.replace("ESK_", "RES_")
            return res_units.get(res_code, "—")
        
        total_no_nds = 0
        total_with_nds = 0
        
        # Записываем данные
        total_no_nds = 0
        total_with_nds = 0
        
        for idx, item in enumerate(items, 1):
            row = idx + 1
            
            price_pu_no = item.price_no_nds or 0
            price_pu_with = item.price_with_nds or 0
            price_va_no = item.price_va_no_nds or 0
            price_va_with = item.price_va_with_nds or 0
            price_truba_no = item.price_truba_no_nds or 0
            price_truba_with = item.price_truba_with_nds or 0
            item_total_no = price_pu_no + price_va_no + price_truba_no
            item_total_with = price_pu_with + price_va_with + price_truba_with
            
            total_no_nds += item_total_no
            total_with_nds += item_total_with
            
            contract_date_str = ""
            plan_date_str = ""
            try:
                contract_date_str = item.contract_date.strftime("%d.%m.%Y") if item.contract_date else ""
            except Exception:
                pass
            try:
                plan_date_str = item.plan_date.strftime("%d.%m.%Y") if item.plan_date else ""
            except Exception:
                pass

            data = [
                idx,
                "Сочинский ПЭС",
                get_res_name(item.current_unit),
                item.consumer or "",
                item.address or "",
                item.contract_number or "",
                contract_date_str,
                plan_date_str,
                item.power if item.power is not None else "",
                item.pu_type or "",
                item.serial_number or "",
                item.faza or "",
                item.work_type_name or "",
                item.lsr_va or "",
                price_va_with,
                "Да" if item.trubostoyka else "Нет",
                item.lsr_truba or "",
                price_truba_with if item.trubostoyka else "",
                item_total_with,
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = thin_border
                
                # Выравнивание
                if col == 1:  # №
                    cell.alignment = center_alignment
                elif col in [7, 8, 9, 12, 16]:  # Даты, мощность, фазность, трубостойка
                    cell.alignment = center_alignment
                elif col in [15, 18, 19]:  # Деньги 
                    cell.alignment = money_alignment
                    if isinstance(value, (int, float)) and value > 0:
                        cell.number_format = '#,##0.00'
                else:
                    cell.alignment = cell_alignment
            
            ws.row_dimensions[row].height = 30
        
        # Итоговая строка
        total_row = len(items) + 2
        ws.cell(row=total_row, column=1, value="ИТОГО:")
        ws.cell(row=total_row, column=1).font = Font(bold=True)
        ws.cell(row=total_row, column=1).alignment = Alignment(horizontal="right")
        
        ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=18)

        ws.cell(row=total_row, column=19, value=total_with_nds)
        ws.cell(row=total_row, column=19).font = Font(bold=True)
        ws.cell(row=total_row, column=19).number_format = '#,##0.00'
        ws.cell(row=total_row, column=19).border = thin_border
        ws.cell(row=total_row, column=19).alignment = money_alignment
        
        # Количество ПУ
        ws.cell(row=total_row + 1, column=1, value=f"Всего ПУ: {len(items)} шт.")
        ws.cell(row=total_row + 1, column=1).font = Font(bold=True)
        
        # Сохраняем в поток
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Безопасное имя файла (ASCII + URL-encoded для UTF-8)
        safe_request_number = re.sub(r'[^\x00-\x7F]', '', request_number.replace("/", "-").replace("\\", "-")).strip("_- ")
        safe_contract = re.sub(r'[^\x00-\x7F]', '', (request_contract or "").replace("/", "-").replace("\\", "-")).strip("_- ")
        filename_ascii = f"Zayavka_{safe_request_number}_{safe_contract}.xlsx".replace(" ", "_")
        filename_utf8 = f"Заявка_{request_number}_{request_contract or ''}.xlsx"
        
        headers = {
            "Content-Disposition": f"attachment; filename=\"{filename_ascii}\"; filename*=UTF-8''{quote(filename_utf8)}"
        }
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers
        )
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"Export error: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(500, f"Ошибка экспорта: {str(e)}")

@app.get("/api/requests/last")
def get_last_request(db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    """Получить последний номер заявки для подсказки"""
    current_year = datetime.utcnow().year
    year_short = str(current_year)[-2:]
    
    last_item = db.query(PUItem).filter(
        PUItem.request_number != None,
        PUItem.request_number != "",
        PUItem.request_number.like(f"%-{year_short}")
    ).order_by(PUItem.id.desc()).first()
    
    next_num = 1
    last_contract = ""
    
    if last_item and last_item.request_number:
        try:
            last_num = int(last_item.request_number.split("-")[0])
            next_num = last_num + 1
            last_contract = last_item.request_contract or ""
        except:
            pass
    
    return {
        "next_number": f"{next_num}-{year_short}",
        "last_contract": last_contract,
        "suggested": f"№ {next_num}-{year_short} Договор № {last_contract}" if last_contract else f"№ {next_num}-{year_short}"
    }

@app.post("/api/requests/create")
def create_request(data: dict, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    """Создать заявку ЭСК"""
    if not is_esk_admin(user) and not is_esk_user(user):
        raise HTTPException(403, "Только ЭСК может формировать заявки")
    
    item_ids = data.get("item_ids", [])
    request_number = data.get("request_number")  # например "1-26"
    request_contract = data.get("request_contract")  # например "147"
    
    if not item_ids:
        raise HTTPException(400, "Не выбраны ПУ")
    if not request_number:
        raise HTTPException(400, "Не указан номер заявки")
    
    # Обновляем ПУ
    for item_id in item_ids:
        item = db.query(PUItem).filter(PUItem.id == item_id).first()
        if item:
            item.request_number = request_number
            item.request_contract = request_contract
            # Копируем work_type_name из ТТР если есть
            if item.ttr_esk_id:
                ttr = db.query(TTR_ESK).filter(TTR_ESK.id == item.ttr_esk_id).first()
                if ttr and ttr.work_type_name:
                    item.work_type_name = ttr.work_type_name
    
    db.commit()
    
    return {
        "created": len(item_ids), 
        "request_number": request_number,
        "request_contract": request_contract,
        "display_name": f"№ {request_number} Договор № {request_contract}" if request_contract else f"№ {request_number}"
    }

@app.post("/api/requests/modify")
def modify_request(data: dict, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    """Добавить/удалить ПУ из заявки (только с паролем)"""
    if data.get("admin_code") != settings.ADMIN_CODE:
        raise HTTPException(403, "Неверный код администратора")
    
    action = data.get("action")  # "add" или "remove"
    item_ids = data.get("item_ids", [])
    request_number = data.get("request_number")
    request_contract = data.get("request_contract")
    
    if action == "add":
        for item_id in item_ids:
            item = db.query(PUItem).filter(PUItem.id == item_id).first()
            if item:
                item.request_number = request_number
                item.request_contract = request_contract
    elif action == "remove":
        for item_id in item_ids:
            item = db.query(PUItem).filter(PUItem.id == item_id).first()
            if item:
                item.request_number = None
                item.request_contract = None
    
    db.commit()
    return {"ok": True, "modified": len(item_ids)}


@app.post("/api/requests/{request_number}/recalculate")
def recalculate_request_prices(request_number: str, data: dict, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    """Пересчитать стоимость выбранных ПУ в заявке по актуальным ценам ТТР ЭСК"""
    if not is_sue_admin(user):
        raise HTTPException(403, "Только СУЭ может пересчитывать стоимости")

    item_ids = data.get("item_ids", [])
    if not item_ids:
        raise HTTPException(400, "Не выбраны ПУ для пересчёта")

    updated = 0
    errors = []

    for item_id in item_ids:
        item = db.query(PUItem).filter(
            PUItem.id == item_id,
            PUItem.request_number == request_number
        ).first()
        if not item:
            errors.append(f"ПУ #{item_id} не найден в заявке")
            continue

        # --- Пересчёт трубостойки ---
        if item.trubostoyka:
            ttr_truba = db.query(TTR_ESK).filter(
                TTR_ESK.ttr_type == "TRUBOSTOYKA",
                TTR_ESK.is_active == True
            ).first()
            if ttr_truba:
                item.lsr_truba = ttr_truba.lsr_number
                item.price_truba_no_nds = ttr_truba.price_no_nds or 0
                item.price_truba_with_nds = ttr_truba.price_with_nds or 0
            else:
                item.lsr_truba = None
                item.price_truba_no_nds = 0
                item.price_truba_with_nds = 0
        else:
            item.lsr_truba = None
            item.price_truba_no_nds = 0
            item.price_truba_with_nds = 0

        # --- Пересчёт ВА (основной ЛСР) ---
        if item.faza and item.form_factor and item.va_type:
            q = db.query(TTR_ESK).filter(
                TTR_ESK.ttr_type == "PU",
                TTR_ESK.faza == item.faza,
                TTR_ESK.form_factor == item.form_factor,
                TTR_ESK.va_type == item.va_type,
                TTR_ESK.is_active == True
            )
            ttr_va = None
            if item.pu_type:
                pu_upper = item.pu_type.upper()
                for t in q.all():
                    if t.pu_pattern and t.pu_pattern.upper() in pu_upper:
                        ttr_va = t
                        break
            if not ttr_va:
                ttr_va = q.first()

            if ttr_va:
                item.lsr_va = ttr_va.lsr_number
                item.price_va_no_nds = ttr_va.price_no_nds or 0
                item.price_va_with_nds = ttr_va.price_with_nds or 0
                item.lsr_number = ttr_va.lsr_number
                item.price_no_nds = ttr_va.price_no_nds or 0
                item.price_with_nds = ttr_va.price_with_nds or 0
                item.ttr_esk_id = ttr_va.id
                item.work_type_name = ttr_va.work_type_name
            else:
                errors.append(f"ПУ #{item_id}: не найден ТТР ЭСК для faza={item.faza}, form_factor={item.form_factor}, va_type={item.va_type}")
        else:
            errors.append(f"ПУ #{item_id}: недостаточно данных для подбора ТТР (faza/form_factor/va_type)")

        updated += 1

    db.commit()
    return {"ok": True, "updated": updated, "errors": errors}


@app.post("/api/requests/{request_number}/remove-items")
def remove_items_from_request(request_number: str, data: dict, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    """Удалить выбранные ПУ из заявки. Если заявка пустая — она автоматически исчезает."""
    if not is_sue_admin(user):
        raise HTTPException(403, "Только СУЭ может изменять заявки")
    if data.get("admin_code") != settings.ADMIN_CODE:
        raise HTTPException(403, "Неверный код администратора")

    item_ids = data.get("item_ids", [])
    if not item_ids:
        raise HTTPException(400, "Не выбраны ПУ для удаления")

    removed = 0
    for item_id in item_ids:
        item = db.query(PUItem).filter(
            PUItem.id == item_id,
            PUItem.request_number == request_number
        ).first()
        if item:
            item.request_number = None
            item.request_contract = None
            removed += 1

    db.commit()

    # Проверяем остались ли ПУ в заявке
    remaining = db.query(PUItem).filter(PUItem.request_number == request_number).count()

    return {"ok": True, "removed": removed, "request_deleted": remaining == 0, "remaining": remaining}

@app.get("/api/requests/search-available")
def search_available_for_request(
    request_number: str,
    request_contract: Optional[str] = None,
    q: str = "",
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user)
):
    """Поиск ПУ для добавления в существующую заявку ЭСК"""
    if not is_sue_admin(user):
        raise HTTPException(403, "Только СУЭ может добавлять ПУ в заявки")
    
    if not q or len(q) < 2:
        return []
    
    # Проверяем что заявка существует
    exists_q = db.query(PUItem).filter(PUItem.request_number == request_number)
    if request_contract:
        exists_q = exists_q.filter(PUItem.request_contract == request_contract)
    if exists_q.count() == 0:
        raise HTTPException(404, "Заявка не найдена или пуста")
    
    # Ищем ПУ: согласованные, без заявки (для ЭСК участок не важен)
    query = db.query(PUItem).options(joinedload(PUItem.current_unit)).filter(
        PUItem.request_number.is_(None),
        PUItem.approval_status == ApprovalStatus.APPROVED,
    )
    
    # Поиск по серийному номеру, договору или потребителю
    search_filter = or_(
        PUItem.serial_number.ilike(f"%{q}%"),
        PUItem.contract_number.ilike(f"%{q}%"),
        PUItem.consumer.ilike(f"%{q}%"),
    )
    query = query.filter(search_filter)
    
    items = query.limit(20).all()
    
    return [{
        "id": i.id,
        "serial_number": i.serial_number,
        "pu_type": i.pu_type,
        "consumer": i.consumer,
        "address": i.address,
        "contract_number": i.contract_number,
        "power": i.power,
        "faza": i.faza,
        "work_type_name": i.work_type_name,
        "price_with_nds": (i.price_truba_with_nds or 0) + (i.price_va_with_nds or 0),
        "current_unit_name": i.current_unit.name if i.current_unit else None,
    } for i in items]

@app.post("/api/requests/{request_number}/add-items")
def add_items_to_request(request_number: str, data: dict, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    """Добавить ПУ в существующую заявку ЭСК — только СУЭ"""
    if not is_sue_admin(user):
        raise HTTPException(403, "Только СУЭ может добавлять ПУ в заявки")
    
    item_ids = data.get("item_ids", [])
    request_contract = data.get("request_contract")
    
    if not item_ids:
        raise HTTPException(400, "Не выбраны ПУ для добавления")
    
    # Проверяем что заявка существует
    existing_count = db.query(PUItem).filter(PUItem.request_number == request_number).count()
    if existing_count == 0:
        raise HTTPException(404, "Заявка не найдена")
    
    added = 0
    for item_id in item_ids:
        item = db.query(PUItem).filter(
            PUItem.id == item_id,
            PUItem.request_number.is_(None),
            PUItem.approval_status == ApprovalStatus.APPROVED,
        ).first()
        if item:
            item.request_number = request_number
            item.request_contract = request_contract
            # Копируем work_type_name из ТТР если есть
            if item.ttr_esk_id:
                ttr = db.query(TTR_ESK).filter(TTR_ESK.id == item.ttr_esk_id).first()
                if ttr and ttr.work_type_name:
                    item.work_type_name = ttr.work_type_name
            added += 1
    
    db.commit()
    
    total = db.query(PUItem).filter(PUItem.request_number == request_number).count()
    
    return {"added": added, "total": total, "request_number": request_number}

@app.get("/api/memo/generate")
def generate_memo(
    tz_number: Optional[str] = None,
    request_number: Optional[str] = None,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user)
):
    """Генерация данных для служебной записки"""
    if not is_sue_admin(user):
        raise HTTPException(403, "Только СУЭ может формировать служебки")
    
    if tz_number:
        items = db.query(PUItem).filter(PUItem.tz_number == tz_number).all()
        doc_type = "ТЗ"
        doc_number = tz_number
    elif request_number:
        items = db.query(PUItem).filter(PUItem.request_number == request_number).all()
        doc_type = "Заявка"
        doc_number = request_number
    else:
        raise HTTPException(400, "Укажите номер ТЗ или заявки")
    
    if not items:
        raise HTTPException(404, "ПУ не найдены")
    
    # Группируем по РЭС/ЭСК
    units_data = {}
    for item in items:
        unit_name = item.current_unit.name if item.current_unit else "Не указано"
        if unit_name not in units_data:
            units_data[unit_name] = []
        units_data[unit_name].append({
            "serial_number": item.serial_number,
            "pu_type": item.pu_type,
            "contract_number": item.contract_number,
            "consumer": item.consumer,
            "address": item.address,
            "power": item.power
        })
    
    return {
        "doc_type": doc_type,
        "doc_number": doc_number,
        "date": datetime.utcnow().strftime("%d.%m.%Y"),
        "total_count": len(items),
        "units": units_data
    }

# ==================== API: ИМПОРТ ДАННЫХ ИЗ EXCEL ====================

@app.post("/api/pu/import-techpris")
async def import_techpris_data(file: UploadFile = File(...), db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    """Импорт данных Техприс по номеру договора"""
    contents = await file.read()
    xl = pd.ExcelFile(io.BytesIO(contents))
    df = pd.read_excel(xl, header=None)
    
    # Ищем заголовки
    header_row = None
    cols = {}
    
    for idx, row in df.iterrows():
        for col_idx, cell in enumerate(row):
            cell_str = str(cell).lower().strip()
            if 'номер договора' in cell_str or 'договор' in cell_str:
                cols['contract'] = col_idx
                header_row = idx
            elif 'потребитель' in cell_str:
                cols['consumer'] = col_idx
            elif 'адрес' in cell_str and 'объект' in cell_str:
                cols['address'] = col_idx
            elif 'pmax' in cell_str or 'мощность' in cell_str:
                cols['power'] = col_idx
            elif 'дата заключения' in cell_str:
                cols['contract_date'] = col_idx
            elif 'планируемая дата' in cell_str or 'дата исполнения' in cell_str:
                cols['plan_date'] = col_idx
        if header_row is not None and len(cols) >= 2:
            break
    
    if 'contract' not in cols:
        raise HTTPException(400, "Не найдена колонка 'Номер договора'")
    
    # Читаем данные после заголовка
    data_rows = df.iloc[header_row + 1:].reset_index(drop=True)
    
    # Строим словарь: номер договора -> данные
    import_data = {}
    for _, row in data_rows.iterrows():
        contract = str(row.iloc[cols['contract']]).strip() if cols.get('contract') is not None else None
        if not contract or contract == 'nan' or len(contract) < 10:
            continue
        
        # Нормализуем формат договора
        contract_clean = re.sub(r'[^\d]', '', contract)
        if len(contract_clean) >= 16:
            contract_formatted = f"{contract_clean[:5]}-{contract_clean[5:7]}-{contract_clean[7:15]}-{contract_clean[15:16]}"
        else:
            contract_formatted = contract
        
        import_data[contract_formatted] = {
            'consumer': str(row.iloc[cols['consumer']]).strip() if cols.get('consumer') is not None else None,
            'address': str(row.iloc[cols['address']]).strip() if cols.get('address') is not None else None,
            'power': row.iloc[cols['power']] if cols.get('power') is not None else None,
            'contract_date': row.iloc[cols['contract_date']] if cols.get('contract_date') is not None else None,
            'plan_date': row.iloc[cols['plan_date']] if cols.get('plan_date') is not None else None,
        }
    
    # Обновляем ПУ
    updated = 0
    items = db.query(PUItem).filter(
        PUItem.status == PUStatus.TECHPRIS,
        PUItem.contract_number != None
    ).all()
    
    for item in items:
        if item.contract_number in import_data:
            data = import_data[item.contract_number]
            if data['consumer'] and data['consumer'] != 'nan':
                item.consumer = data['consumer']
            if data['address'] and data['address'] != 'nan':
                item.address = data['address']
            if data['power'] and str(data['power']) != 'nan':
                try:
                    item.power = float(data['power'])
                except:
                    pass
            if data['contract_date'] and str(data['contract_date']) != 'nan':
                try:
                    if isinstance(data['contract_date'], pd.Timestamp):
                        item.contract_date = data['contract_date'].date()
                    elif isinstance(data['contract_date'], datetime):
                        item.contract_date = data['contract_date'].date()
                except:
                    pass
            if data['plan_date'] and str(data['plan_date']) != 'nan':
                try:
                    if isinstance(data['plan_date'], pd.Timestamp):
                        item.plan_date = data['plan_date'].date()
                    elif isinstance(data['plan_date'], datetime):
                        item.plan_date = data['plan_date'].date()
                except:
                    pass
            updated += 1
    
    db.commit()
    return {"updated": updated, "total_in_file": len(import_data)}


@app.post("/api/pu/import-zamena")
async def import_zamena_data(file: UploadFile = File(...), db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    """Импорт данных Замена/ИЖЦ по номеру счётчика"""
    contents = await file.read()
    xl = pd.ExcelFile(io.BytesIO(contents))
    df = pd.read_excel(xl, header=None)
    
    # Ищем заголовки
    cols = {}
    header_row = None
    
    for idx, row in df.iterrows():
        for col_idx, cell in enumerate(row):
            cell_str = str(cell).lower().strip()
            if 'номер счетчика' in cell_str or 'номер пу' in cell_str or 'заводской' in cell_str:
                cols['serial'] = col_idx
                header_row = idx
            elif 'лс' in cell_str or 'лицевой' in cell_str:
                cols['ls'] = col_idx
        if header_row is not None and 'serial' in cols:
            break
    
    if 'serial' not in cols:
        raise HTTPException(400, "Не найдена колонка 'Номер счетчика'")
    if 'ls' not in cols:
        raise HTTPException(400, "Не найдена колонка 'ЛС'")
    
    # Читаем данные
    data_rows = df.iloc[header_row + 1:].reset_index(drop=True)
    
    # Строим словарь: номер счётчика -> ЛС
    import_data = {}
    for _, row in data_rows.iterrows():
        serial = str(row.iloc[cols['serial']]).strip()
        ls = str(row.iloc[cols['ls']]).strip()
        if serial and serial != 'nan' and ls and ls != 'nan':
            import_data[serial] = ls
    
    # Обновляем ПУ
    updated = 0
    items = db.query(PUItem).filter(
        PUItem.status.in_([PUStatus.ZAMENA, PUStatus.IZHC])
    ).all()
    
    for item in items:
        if item.serial_number in import_data:
            item.ls_number = import_data[item.serial_number]
            updated += 1
    
    db.commit()
    return {"updated": updated, "total_in_file": len(import_data)}


@app.post("/api/pu/import-lookup-techpris")
async def import_lookup_techpris(
    file: UploadFile = File(...),
    contract_number: str = Form(...),
    current_user: User = Depends(get_current_user)
):
    """Поиск данных по номеру договора в Excel файле"""
    content = await file.read()
    df = pd.read_excel(io.BytesIO(content), header=None)
    
    # Ищем строку с заголовками
    header_row = None
    for idx, row in df.iterrows():
        for cell in row.values:
            if 'номер договора' in str(cell).lower():
                header_row = idx
                break
        if header_row is not None:
            break
    
    if header_row is None:
        return {"found": False, "error": "Заголовок не найден"}
    
    # Переименовываем колонки
    df.columns = df.iloc[header_row]
    df = df.iloc[header_row + 1:].reset_index(drop=True)
    
    # Ищем нужные колонки
    col_map = {}
    for col in df.columns:
        col_str = str(col).strip().lower() if pd.notna(col) else ''
        if 'номер договора' in col_str:
            col_map['contract'] = col
        elif 'потребитель' in col_str:
            col_map['consumer'] = col
        elif 'адрес' in col_str:
            col_map['address'] = col
        elif 'pmax' in col_str:
            col_map['power'] = col
        elif 'p(запраш' in col_str:
            col_map['power_req'] = col
        elif 'дата заключения' in col_str:
            col_map['contract_date'] = col
        elif 'планируемая дата' in col_str:
            col_map['plan_date'] = col
    
    if 'contract' not in col_map:
        return {"found": False}
    
    # Нормализуем номер договора
    contract_clean = contract_number.replace('-', '').replace(' ', '').lower()
    
    for idx, row in df.iterrows():
        cell_value = str(row.get(col_map['contract'], '')).replace('-', '').replace(' ', '').lower()
        if not cell_value or cell_value == 'nan' or len(cell_value) < 10:
            continue
        
        if contract_clean == cell_value or contract_clean in cell_value:
            result = {"found": True}
            
            if 'consumer' in col_map:
                val = row.get(col_map['consumer'])
                if pd.notna(val) and str(val) != 'nan':
                    result['consumer'] = str(val).strip()
            
            if 'address' in col_map:
                val = row.get(col_map['address'])
                if pd.notna(val) and str(val) != 'nan':
                    result['address'] = str(val).strip()
            
            power_col = col_map.get('power_req') or col_map.get('power')
            if power_col:
                val = row.get(power_col)
                if pd.notna(val) and str(val) != 'nan':
                    try:
                        result['power'] = float(val)
                    except:
                        pass
            
            if 'contract_date' in col_map:
                val = row.get(col_map['contract_date'])
                if pd.notna(val) and str(val) != 'nan':
                    try:
                        if hasattr(val, 'strftime'):
                            result['contract_date'] = val.strftime('%Y-%m-%d')
                        else:
                            result['contract_date'] = str(val)[:10]
                    except:
                        pass
            
            if 'plan_date' in col_map:
                val = row.get(col_map['plan_date'])
                if pd.notna(val) and str(val) != 'nan':
                    try:
                        if hasattr(val, 'strftime'):
                            result['plan_date'] = val.strftime('%Y-%m-%d')
                        else:
                            result['plan_date'] = str(val)[:10]
                    except:
                        pass
            
            return result
    
    return {"found": False}


@app.post("/api/pu/import-lookup-zamena")
async def import_lookup_zamena(
    file: UploadFile = File(...),
    serial_number: str = Form(...),
    current_user: User = Depends(get_current_user)
):
    """Поиск ЛС по серийному номеру счётчика в выгрузке 1С"""
    content = await file.read()
    df = pd.read_excel(io.BytesIO(content), header=None)
    
    # Ищем колонки с заголовками "Номер счетчика" и "ЛС / ЛС СТЕК"
    col_serial = None
    col_ls = None
    header_row = None
    
    # Проходим первые 10 строк в поисках заголовков
    for idx in range(min(10, len(df))):
        row = df.iloc[idx]
        for col_idx, cell in enumerate(row.values):
            cell_str = str(cell).lower().strip() if pd.notna(cell) else ''
            if 'номер счетчика' in cell_str or 'номер счётчика' in cell_str:
                col_serial = col_idx
                header_row = idx
            elif 'лс' in cell_str and ('стек' in cell_str or col_idx < 10):
                col_ls = col_idx
    
    if col_serial is None:
        return {"found": False, "error": "Колонка 'Номер счетчика' не найдена"}
    if col_ls is None:
        return {"found": False, "error": "Колонка 'ЛС / ЛС СТЕК' не найдена"}
    
    # Нормализуем серийный номер для поиска
    serial_clean = serial_number.strip().lower()
    
    # Ищем в данных (после заголовка)
    for idx in range(header_row + 1, len(df)):
        row = df.iloc[idx]
        cell_value = str(row.iloc[col_serial]).strip().lower() if pd.notna(row.iloc[col_serial]) else ''
        
        # Пропускаем пустые
        if not cell_value or cell_value == 'nan':
            continue
        
        # Сравниваем (точное совпадение или содержит)
        if serial_clean == cell_value or serial_clean in cell_value or cell_value in serial_clean:
            ls_val = row.iloc[col_ls]
            if pd.notna(ls_val) and str(ls_val) != 'nan':
                return {"found": True, "ls_number": str(ls_val).strip()}
    
    return {"found": False, "error": f"Счётчик {serial_number} не найден в файле"}

# ==================== ИНИЦИАЛИЗАЦИЯ БД ====================
def init_db():
    db = SessionLocal()
    
    # Роли
    roles = {}
    for name, code in [
        ("СУЭ Администратор", RoleCode.SUE_ADMIN),
        ("Лаборатория", RoleCode.LAB_USER),
        ("ЭСК Администратор", RoleCode.ESK_ADMIN),
        ("Пользователь РЭС", RoleCode.RES_USER),
        ("Пользователь ЭСК", RoleCode.ESK_USER),
        ("ОКС Администратор", RoleCode.OKS_ADMIN),
        ("Пользователь ОКС", RoleCode.OKS_USER),
    ]:
        r = db.query(Role).filter(Role.code == code).first()
        if not r:
            r = Role(name=name, code=code)
            db.add(r)
            db.flush()
        roles[code] = r
    
    # Подразделения
    units = {}
    
    # Первый уровень
    for name, code, utype in [
        ("Служба учета электроэнергии", "SUE", UnitType.SUE),
        ("Лаборатория", "LAB", UnitType.LAB),
        ("ЭСК", "ESK", UnitType.ESK),
        ("ОКС", "OKS", UnitType.OKS),
    ]:
        u = db.query(Unit).filter(Unit.code == code).first()
        if not u:
            u = Unit(name=name, code=code, unit_type=utype)
            db.add(u)
            db.flush()
        units[code] = u
    
    # 7 РЭС + 7 ЭСК подразделений
    res_data = [
        ("Адлерский РЭС", "RES_ADLER", "а"),
        ("Дагомысский РЭС", "RES_DAGOMYS", "д"),
        ("Краснополянский РЭС", "RES_KRASNAYA", "к"),
        ("Лазаревский РЭС", "RES_LAZAREV", "л"),
        ("Сочинский РЭС", "RES_SOCHI", "с"),
        ("Туапсинский РЭС", "RES_TUAPSE", "т"),
        ("Хостинский РЭС", "RES_HOSTA", "х"),
    ]
    
    for res_name, res_code, short in res_data:
        res = db.query(Unit).filter(Unit.code == res_code).first()
        if not res:
            res = Unit(name=res_name, code=res_code, unit_type=UnitType.RES, short_code=short)
            db.add(res)
            db.flush()
        
        # Подразделение ЭСК
        esk_code = res_code.replace("RES_", "ESK_")
        esk_name = res_name.replace(" РЭС", " ЭСК")
        esk = db.query(Unit).filter(Unit.code == esk_code).first()
        if not esk:
            esk = Unit(name=esk_name, code=esk_code, unit_type=UnitType.ESK_UNIT, short_code=short, parent_id=units["ESK"].id)
            db.add(esk)
        
        # Участок ОКС (например "ОКС Адлерский РЭС")
        oks_code = res_code.replace("RES_", "OKS_")
        oks_name = f"ОКС {res_name}"
        oks = db.query(Unit).filter(Unit.code == oks_code).first()
        if not oks:
            oks = Unit(name=oks_name, code=oks_code, unit_type=UnitType.OKS_UNIT, short_code=short, parent_id=units["OKS"].id)
            db.add(oks)
    
    # Тестовые пользователи
    if not db.query(User).filter(User.username == "admin").first():
        db.add(User(username="admin", password_hash=hash_password("admin123"), full_name="Администратор СУЭ", role_id=roles[RoleCode.SUE_ADMIN].id, unit_id=units["SUE"].id))
    if not db.query(User).filter(User.username == "lab").first():
        db.add(User(username="lab", password_hash=hash_password("lab123"), full_name="Оператор Лаборатории", role_id=roles[RoleCode.LAB_USER].id, unit_id=units["LAB"].id))
    if not db.query(User).filter(User.username == "esk").first():
        db.add(User(username="esk", password_hash=hash_password("esk123"), full_name="Администратор ЭСК", role_id=roles[RoleCode.ESK_ADMIN].id, unit_id=units["ESK"].id))
    if not db.query(User).filter(User.username == "oks").first():
        db.add(User(username="oks", password_hash=hash_password("oks123"), full_name="Администратор ОКС", role_id=roles[RoleCode.OKS_ADMIN].id, unit_id=units["OKS"].id))
    
    # Тестовые ТТР для РЭС
    for i in range(1, 8):
        for ttr_type, prefix in [("OU", "ОУ"), ("OL", "ОЛ"), ("OR", "ОР")]:
            code = f"ТТР-{i} {prefix}"
            if not db.query(TTR_RES).filter(TTR_RES.code == code).first():
                db.add(TTR_RES(code=code, name=f"Типовое решение {prefix} #{i}", ttr_type=ttr_type))
    
    # Тестовые ТТР для ЭСК (по комбинациям параметров)

    # (ttr_type, pu_pattern, faza, form_factor, va_type, lsr, price_no, price_with)
    ttr_esk_data = [
        # ПУ
        ("PU", "НАРТИС", "1ф", "split", "opora", "ЛСР-001", 5000, 6000),
        ("PU", "НАРТИС", "1ф", "split", "fasad", "ЛСР-002", 5500, 6600),
        ("PU", "НАРТИС", "1ф", "classic", "opora", "ЛСР-003", 4500, 5400),
        ("PU", "НАРТИС", "3ф", "split", "opora", "ЛСР-004", 7000, 8400),
        ("PU", "НАРТИС", "3ф", "classic", "fasad", "ЛСР-005", 6500, 7800),
        # Трубостойки
        ("TRUBOSTOYKA", None, None, None, None, "ЛСР-Т01", 8000, 9600),
        ("TRUBOSTOYKA", None, None, None, None, "ЛСР-Т02", 10000, 12000),
        # Ответвления
        ("OTVETVLENIE", None, None, None, None, "ЛСР-О01", 3000, 3600),
    ]
    for ttr_type, pu_pattern, faza, form_factor, va_type, lsr, price_no, price_with in ttr_esk_data:
        existing = db.query(TTR_ESK).filter(TTR_ESK.lsr_number == lsr).first()
        if not existing:
            db.add(TTR_ESK(
                ttr_type=ttr_type,
                pu_pattern=pu_pattern,
                faza=faza,
                form_factor=form_factor,
                va_type=va_type,
                lsr_number=lsr,
                price_no_nds=price_no,
                price_with_nds=price_with
            ))
    
    db.commit()
    db.close()
    print("✅ БД инициализирована!")

# ==================== АВТОМИГРАЦИЯ СХЕМЫ БД ====================
def ensure_db_schema():
    """Добавляет новые таблицы и колонки без удаления данных"""
    from sqlalchemy import inspect, text
    
    db = SessionLocal()
    inspector = inspect(engine)
    existing_tables = inspector.get_table_names()
    
    try:
        # 1. Создаём недостающие таблицы
        Base.metadata.create_all(bind=engine)
        print("✅ Таблицы проверены")
        
        # 1.5 Добавляем новые значения в существующие PostgreSQL enum-типы.
        # Это НЕ покрывается create_all: для уже созданного native enum-типа
        # новые значения нужно добавлять через ALTER TYPE, иначе вставка
        # подразделений/ролей ОКС упадёт с "invalid input value for enum".
        enum_value_additions = {
            "unittype": ["OKS", "OKS_UNIT"],
            "rolecode": ["OKS_ADMIN", "OKS_USER"],
        }
        try:
            with engine.connect() as conn:
                conn = conn.execution_options(isolation_level="AUTOCOMMIT")
                existing_enum_types = [
                    r[0] for r in conn.execute(text(
                        "SELECT typname FROM pg_type WHERE typtype = 'e'"
                    ))
                ]
                for type_name, values in enum_value_additions.items():
                    # Если тип ещё не создан (свежая БД) — create_all уже
                    # создал его со всеми актуальными значениями, ничего не делаем.
                    if type_name not in existing_enum_types:
                        continue
                    for val in values:
                        try:
                            conn.execute(text(
                                f"ALTER TYPE {type_name} ADD VALUE IF NOT EXISTS '{val}'"
                            ))
                            print(f"  ➕ enum {type_name} += {val}")
                        except Exception as e:
                            msg = str(e).lower()
                            if 'already exists' not in msg and 'duplicate' not in msg:
                                print(f"  ⚠️ enum {type_name} {val}: {e}")
        except Exception as e:
            # Не для PostgreSQL (например SQLite) — enum-типов нет, пропускаем
            print(f"  ℹ️ Пропуск миграции enum: {e}")
        
        # 2. Добавляем недостающие колонки в существующие таблицы
        for table_name, table in Base.metadata.tables.items():
            if table_name not in existing_tables:
                continue
                
            existing_columns = [c['name'] for c in inspector.get_columns(table_name)]
            
            for column in table.columns:
                if column.name not in existing_columns:
                    # Определяем SQL тип
                    col_type = str(column.type)
                    
                    # Для PostgreSQL enum нужна особая обработка
                    if 'VARCHAR' in col_type.upper():
                        sql_type = col_type
                    elif 'INTEGER' in col_type.upper():
                        sql_type = 'INTEGER'
                    elif 'FLOAT' in col_type.upper() or 'DOUBLE' in col_type.upper():
                        sql_type = 'FLOAT'
                    elif 'BOOLEAN' in col_type.upper():
                        sql_type = 'BOOLEAN'
                    elif 'DATE' == col_type.upper():
                        sql_type = 'DATE'
                    elif 'DATETIME' in col_type.upper() or 'TIMESTAMP' in col_type.upper():
                        sql_type = 'TIMESTAMP'
                    elif 'TEXT' in col_type.upper():
                        sql_type = 'TEXT'
                    else:
                        sql_type = 'VARCHAR(255)'  # fallback
                    
                    sql = f'ALTER TABLE "{table_name}" ADD COLUMN "{column.name}" {sql_type} NULL'
                    
                    try:
                        db.execute(text(sql))
                        db.commit()
                        print(f"  ➕ Добавлена колонка: {table_name}.{column.name} ({sql_type})")
                    except Exception as e:
                        db.rollback()
                        # Игнорируем если колонка уже есть
                        if 'already exists' not in str(e).lower() and 'duplicate' not in str(e).lower():
                            print(f"  ⚠️ Ошибка {table_name}.{column.name}: {e}")
        
        # 3. Индексы на часто фильтруемых колонках (ускоряют списки/дашборд/аналитику/согласование)
        index_defs = [
            ("ix_pu_items_current_unit_id", "pu_items", "current_unit_id"),
            ("ix_pu_items_status", "pu_items", "status"),
            ("ix_pu_items_approval_status", "pu_items", "approval_status"),
            ("ix_pu_items_tz_number", "pu_items", "tz_number"),
            ("ix_pu_items_request_number", "pu_items", "request_number"),
            ("ix_pu_items_contract_number", "pu_items", "contract_number"),
            ("ix_pu_items_register_id", "pu_items", "register_id"),
            ("ix_pu_items_target_unit_id", "pu_items", "target_unit_id"),
            ("ix_pu_materials_pu_item_id", "pu_materials", "pu_item_id"),
            ("ix_pu_movements_pu_item_id", "pu_movements", "pu_item_id"),
        ]
        for idx_name, tbl, col in index_defs:
            if tbl not in existing_tables:
                continue
            try:
                db.execute(text(f'CREATE INDEX IF NOT EXISTS {idx_name} ON "{tbl}" ("{col}")'))
                db.commit()
            except Exception as e:
                db.rollback()
                if 'already exists' not in str(e).lower():
                    print(f"  ⚠️ Индекс {idx_name}: {e}")
        print("✅ Индексы проверены")
        
        print("✅ Схема БД актуальна")
        
    finally:
        db.close()

ensure_db_schema()
init_db()
